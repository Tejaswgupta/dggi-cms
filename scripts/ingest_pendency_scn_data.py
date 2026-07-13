"""
Ingest Pendency / SCN / NON-IR Excel data into Supabase.

Source: 'Pendency, SCN & Saadhit Target-2 Data.xlsx'
  - Sheet 'FInal Pendency' → dggi_records (is_ir=True)  ~92 rows
  - Sheet 'SCN'             → dggi_scn_records            ~18 rows
  - Sheet 'NON-IR'          → dggi_records (is_ir=False) ~32 rows

Upsert strategy:
  - dggi_records:     match on (workspace_id, file_no); update if exists, insert if not
  - dggi_scn_records: match on (workspace_id, scn_no);  update if exists, insert if not

Usage:
    python3 scripts/ingest_pendency_scn_data.py [/path/to/file.xlsx] [--dry-run]
"""

import csv
import os
import sys
from datetime import date, datetime

import openpyxl
from supabase import create_client

SUPABASE_URL = "https://zrkvvedwycdcjjheewef.supabase.co"
SERVICE_ROLE_KEY = "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6Inpya3Z2ZWR3eWNkY2pqaGVld2VmIiwicm9sZSI6InNlcnZpY2Vfcm9sZSIsImlhdCI6MTcwMzAxNjg1NCwiZXhwIjoyMDE4NTkyODU0fQ.ZYgzzv6E--3v2un2uN0jXwHnBvCf0EjPJlGoCQwiqKE"

WORKSPACE_OWNER_EMAIL = "ajinkya.k1@gov.in"

DEFAULT_EXCEL_PATH = os.path.join(
    os.path.expanduser("~"),
    "Downloads",
    "Pendency, SCN & Saadhit Target-2 Data.xlsx",
)

SKIPPED_CSV = os.path.join(os.path.dirname(__file__), "ingest_skipped.csv")


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def current_fy() -> str:
    """Returns current fiscal year in YY-YY format (e.g., '26-27' for FY 2026-2027)."""
    now = datetime.now()
    yr = now.year
    start = yr if now.month >= 4 else yr - 1
    return f"{str(start)[2:]}-{str(start + 1)[2:]}"


def generate_record_id(sb, workspace_id: str, is_ir: bool) -> str:
    """Generate sequential record_id matching the DGGI Excel convention.
    IR cases    → "{seq}/GST/{YYYY-YYYY+1}"  e.g. "001/GST/2026-27"
    NON-IR cases → "NIR-{seq}-{YY-YY}"       e.g. "NIR-001-26-27"
    """
    res = (
        sb.table("dggi_records")
        .select("*", count="exact", head=True)
        .eq("workspace_id", workspace_id)
        .eq("is_ir", is_ir)
        .execute()
    )
    count = res.count or 0
    seq = str(count + 1).zfill(3)
    if is_ir:
        now = datetime.now()
        yr = now.year if now.month >= 4 else now.year - 1
        fy_full = f"{yr}-{str(yr + 1)[2:]}"
        return f"{seq}/GST/{fy_full}"
    return f"NIR-{seq}-{current_fy()}"


def parse_date(val) -> str | None:
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.date().isoformat()
    if isinstance(val, date):
        return val.isoformat()
    s = str(val).strip()
    if not s:
        return None
    for fmt in ("%d.%m.%Y", "%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"):
        try:
            return datetime.strptime(s, fmt).date().isoformat()
        except ValueError:
            pass
    return None


def normalize_group(val) -> str | None:
    if val is None:
        return None
    s = str(val).strip()
    if s.startswith("Group "):
        return s
    if len(s) == 1 and s.upper() in "ABCDEF":
        return f"Group {s.upper()}"
    return None


def clean(val) -> str | None:
    if val is None:
        return None
    s = str(val).strip()
    return s if s else None


def str_amount(val) -> str | None:
    if val is None:
        return None
    try:
        f = float(val)
        return str(f) if f != 0 else None
    except (ValueError, TypeError):
        s = str(val).strip()
        return s if s else None


# ---------------------------------------------------------------------------
# Dry-run helper — checks DB without writing
# ---------------------------------------------------------------------------

def _would_update(sb, workspace_id: str, table: str, key_col: str, key_val: str | None) -> bool:
    if not key_val:
        return False
    res = sb.table(table).select("id").eq("workspace_id", workspace_id).eq(key_col, key_val).execute()
    return bool(res.data)


# ---------------------------------------------------------------------------
# Upsert helpers
# ---------------------------------------------------------------------------

def upsert_dggi_record(sb, workspace_id: str, file_no: str | None, payload: dict, skipped: list, label: str, dry_run: bool = False):
    try:
        if file_no:
            existing = (
                sb.table("dggi_records")
                .select("id,record_id")
                .eq("workspace_id", workspace_id)
                .eq("file_no", file_no)
                .execute()
            )
            if existing.data:
                # Update existing — never overwrite record_id
                if not dry_run:
                    sb.table("dggi_records").update(payload).eq("id", existing.data[0]["id"]).execute()
                return "updated"
        # New insert — always generate record_id
        is_ir = payload.get("is_ir", True)
        full_payload = {
            **payload,
            "workspace_id": workspace_id,
            "file_no": file_no,
            "record_id": generate_record_id(sb, workspace_id, is_ir),
        }
        if not dry_run:
            sb.table("dggi_records").insert(full_payload).execute()
        return "inserted"
    except Exception as e:
        skipped.append({"sheet": label, "file_no": file_no or "", "reason": str(e)})
        return "skipped"


def upsert_scn_record(sb, workspace_id: str, scn_no: str | None, payload: dict, skipped: list, dry_run: bool = False):
    try:
        if scn_no:
            existing = (
                sb.table("dggi_scn_records")
                .select("id")
                .eq("workspace_id", workspace_id)
                .eq("scn_no", scn_no)
                .execute()
            )
            if existing.data:
                if not dry_run:
                    sb.table("dggi_scn_records").update(payload).eq("id", existing.data[0]["id"]).execute()
                return "updated"
        if not dry_run:
            sb.table("dggi_scn_records").insert({**payload, "workspace_id": workspace_id, "scn_no": scn_no}).execute()
        return "inserted"
    except Exception as e:
        skipped.append({"sheet": "SCN", "scn_no": scn_no or "", "reason": str(e)})
        return "skipped"


# ---------------------------------------------------------------------------
# Column mapping report
# ---------------------------------------------------------------------------

# Each entry: (excel_col_index, excel_header, db_column, note)
# note=None means direct map; note="combined" means merged into another field; note="derived" means computed
PENDENCY_MAPPING = [
    (0,  "Sr. No.",                          None,                    "skipped — row counter"),
    (1,  "Name of the Taxpayer",             "taxpayer_name",         None),
    (2,  "GSTIN/PAN",                        "gstins",                None),
    (3,  "IR No./335-J No.",                 "file_no",               "upsert key"),
    (4,  "Date of Detection/IR Date",        "date_of_ir + date_of_initiation", "combined"),
    (5,  "(unlabeled — as-of date)",         None,                    "skipped — static reference date"),
    (6,  "Pending since (days)",             None,                    "skipped — computed value"),
    (7,  "Pendency Year Wise",               None,                    "skipped — derived bucket"),
    (8,  "Detection (in Lakhs)",             "detection_amount",      None),
    (9,  "Additional Detection (in Lakhs)",  None,                    "skipped — no target column"),
    (10, "Recovery (in Lakh)",               "recovery_itc",          None),
    (11, "Additional Recovery (in Lakhs)",   None,                    "skipped — no target column"),
    (12, "Type of the Case",                 "issue_involved",        "combined into issue_involved"),
    (13, "Brief Facts of the Case",          "issue_involved",        "primary part of issue_involved"),
    (14, "Present Status",                   "latest_status",         None),
    (15, "Expected Date of Closure/SCN",     "due_date",              None),
    (16, "Name of SIO",                      None,                    "skipped — no user mapping available"),
    (17, "Group",                            "group",                 "A→Group A"),
    (18, "Case booked under section",        "issue_involved",        "combined into issue_involved"),
    (19, "Whether in DIGIT (DIGIT No.)",     "digit_id",              None),
]

SCN_MAPPING = [
    (0,  "Sr. No.",                          None,                    "skipped — row counter"),
    (1,  "Date of detection",                "remarks",               "combined into remarks"),
    (2,  "Name of the party",               "noticee_name",          None),
    (3,  "GSTIN",                            "gstin_pan",             None),
    (4,  "SCN No.",                          "scn_no",                "upsert key"),
    (5,  "SCN date",                         "date_of_scn",           None),
    (6,  "Amount Involved (in Lakhs)",       "demand_tax",            None),
    (7,  "Name & Designation of AA",         "adjudicating_authority",None),
    (8,  "Adjudicating Authority Single/Common", "common_adjudicating_authority", None),
    (9,  "Adjudicating Commissionerate",     "adjudication_formation",None),
    (10, "Adjudicating Commissionerate Zone","remarks",               "combined into remarks"),
    (11, "Period Involved",                  "period_involved",       None),
    (12, "SIO",                              "sio_name + created_by_name", "name stored as text; no UUID mapping"),
    (13, "Group",                            "group",                 "A→Group A"),
    (14, "Whether updated in DIGIT",         None,                    "skipped — informational only"),
    (15, "MPR Month",                        None,                    "skipped — use dggi_mpr_records instead"),
]

NON_IR_MAPPING = [
    (0,  "Sr. No.",                          None,                    "skipped — row counter"),
    (1,  "Name of GSTIN",                    "taxpayer_name",         None),
    (2,  "GSTIN",                            "gstins",                "also fallback upsert key"),
    (3,  "Initiation date",                  "date_of_initiation + date_of_non_ir", "combined"),
    (4,  "SIO",                              None,                    "skipped — no user mapping available"),
    (5,  "Group",                            "group",                 "A→Group A"),
    (6,  "Mode",                             "mode_of_initiation",    None),
    (7,  "Month (as-of)",                    None,                    "skipped — static reference string"),
    (8,  "Current Status",                   "latest_status",         None),
    (9,  "IR No. (if IR issued)",            "file_no",               "primary upsert key"),
]


def print_mapping_report():
    sheets = [
        ("FInal Pendency", PENDENCY_MAPPING),
        ("SCN",            SCN_MAPPING),
        ("NON-IR",         NON_IR_MAPPING),
    ]
    total_mapped = total_skipped = 0
    print("=" * 70)
    print("COLUMN MAPPING REPORT")
    print("=" * 70)
    for sheet_name, mapping in sheets:
        mapped   = [(i, h, db, n) for i, h, db, n in mapping if db is not None]
        skipped  = [(i, h, db, n) for i, h, db, n in mapping if db is None]
        print(f"\n── {sheet_name}  ({len(mapping)} cols total | {len(mapped)} mapped | {len(skipped)} not mapped) ──")
        print(f"  {'COL':<4} {'EXCEL HEADER':<42} {'→ DB COLUMN':<38} NOTE")
        print(f"  {'-'*4} {'-'*42} {'-'*38} {'-'*20}")
        for i, h, db, note in mapping:
            marker = "✓" if db else "✗"
            db_str = db if db else "(not ingested)"
            note_str = note or ""
            print(f"  {marker} {i:<3} {h[:41]:<42} {db_str[:37]:<38} {note_str}")
        total_mapped   += len(mapped)
        total_skipped  += len(skipped)
    print(f"\n{'=' * 70}")
    print(f"TOTAL  {total_mapped + total_skipped} Excel columns across all sheets")
    print(f"  Mapped   : {total_mapped}")
    print(f"  Not mapped: {total_skipped}")
    print("=" * 70)
    print()


# ---------------------------------------------------------------------------
# Sheet processors
# ---------------------------------------------------------------------------

def process_pendency(ws, sb, workspace_id: str, skipped: list, dry_run: bool = False):
    rows = list(ws.iter_rows(values_only=True))
    inserted = updated = skipped_count = 0

    for row in rows[2:]:  # row 0 = title, row 1 = headers
        taxpayer_name = clean(row[1])
        if not taxpayer_name:
            continue

        detection_date = parse_date(row[4])
        issue_parts = []
        if clean(row[13]):
            issue_parts.append(clean(row[13]))
        if clean(row[12]):
            issue_parts.append(f"Type: {clean(row[12])}")
        if clean(row[18]):
            issue_parts.append(f"Section: {clean(row[18])}")

        payload = {
            "taxpayer_name": taxpayer_name,
            "gstins": clean(row[2]),
            "date_of_ir": detection_date,
            "date_of_initiation": detection_date,
            "detection_amount": str_amount(row[8]),
            "recovery_itc": str_amount(row[10]),
            "issue_involved": " | ".join(issue_parts) if issue_parts else None,
            "latest_status": clean(row[14]),
            "due_date": parse_date(row[15]),
            "group": normalize_group(row[17]),
            "digit_id": clean(row[19]),
            "is_ir": True,
        }
        # Skip handling_io_sio_name — no way to map Excel names to user UUIDs
        payload = {k: v for k, v in payload.items() if v is not None}

        file_no = clean(row[3])
        if dry_run:
            print(f"  [IR] {'UPDATE' if _would_update(sb, workspace_id, 'dggi_records', 'file_no', file_no) else 'INSERT'} | file_no={file_no} | taxpayer={taxpayer_name} | group={payload.get('group')} | date={detection_date}")
        result = upsert_dggi_record(sb, workspace_id, file_no, payload, skipped, "FInal Pendency", dry_run)
        if result == "inserted":
            inserted += 1
        elif result == "updated":
            updated += 1
        else:
            skipped_count += 1

    print(f"[IR Pendency]  {'Would insert' if dry_run else 'Inserted'}: {inserted} | {'Would update' if dry_run else 'Updated'}: {updated} | Skipped: {skipped_count}")


def process_scn(ws, sb, workspace_id: str, skipped: list, dry_run: bool = False):
    rows = list(ws.iter_rows(values_only=True))
    inserted = updated = skipped_count = 0

    for row in rows[2:]:
        noticee_name = clean(row[2])
        if not noticee_name:
            continue

        remarks_parts = []
        if clean(row[10]):
            remarks_parts.append(f"Zone: {clean(row[10])}")
        if clean(row[1]):
            remarks_parts.append(f"Detection: {clean(row[1])}")

        payload = {
            "noticee_name": noticee_name,
            "gstin_pan": clean(row[3]),
            "date_of_scn": parse_date(row[5]),
            "demand_tax": str_amount(row[6]),
            "adjudicating_authority": clean(row[7]),
            "common_adjudicating_authority": clean(row[8]),
            "adjudication_formation": clean(row[9]),
            "period_involved": clean(row[11]),
            "group": normalize_group(row[13]),
            "remarks": " | ".join(remarks_parts) if remarks_parts else None,
            # competency is required for the SCN Register tabs in the UI;
            # Excel has no competency column so default to SIO Competency
            "competency": "SIO Competency",
            "sio_name": clean(row[12]),
            "created_by_name": clean(row[12]),
        }
        payload = {k: v for k, v in payload.items() if v is not None}

        scn_no = clean(row[4])
        if dry_run:
            print(f"  [SCN] {'UPDATE' if _would_update(sb, workspace_id, 'dggi_scn_records', 'scn_no', scn_no) else 'INSERT'} | scn_no={scn_no} | noticee={noticee_name} | group={payload.get('group')} | date_of_scn={payload.get('date_of_scn')}")
        result = upsert_scn_record(sb, workspace_id, scn_no, payload, skipped, dry_run)
        if result == "inserted":
            inserted += 1
        elif result == "updated":
            updated += 1
        else:
            skipped_count += 1

    print(f"[SCN]          {'Would insert' if dry_run else 'Inserted'}: {inserted} | {'Would update' if dry_run else 'Updated'}: {updated} | Skipped: {skipped_count}")


def process_non_ir(ws, sb, workspace_id: str, skipped: list, dry_run: bool = False):
    rows = list(ws.iter_rows(values_only=True))
    inserted = updated = skipped_count = 0

    for row in rows[2:]:
        taxpayer_name = clean(row[1])
        if not taxpayer_name:
            continue

        initiation_date = parse_date(row[3])
        mode = clean(row[6])
        # mode_of_initiation CHECK constraint: Letter/Email/Summons/Inspection/Search
        if mode and mode not in ("Letter", "Email", "Summons", "Inspection", "Search"):
            mode = None

        payload = {
            "taxpayer_name": taxpayer_name,
            "gstins": clean(row[2]),
            "date_of_initiation": initiation_date,
            "date_of_non_ir": initiation_date,
            "group": normalize_group(row[5]),
            "mode_of_initiation": mode,
            "latest_status": clean(row[8]),
            "is_ir": False,
        }
        # Skip handling_io_sio_name — no way to map Excel names to user UUIDs
        payload = {k: v for k, v in payload.items() if v is not None}

        ir_no = clean(row[9])
        gstin = clean(row[2])
        file_no = ir_no or gstin

        if dry_run:
            print(f"  [NON-IR] {'UPDATE' if _would_update(sb, workspace_id, 'dggi_records', 'file_no', file_no) else 'INSERT'} | file_no={file_no} | taxpayer={taxpayer_name} | group={payload.get('group')} | mode={mode} | date={initiation_date}")
        result = upsert_dggi_record(sb, workspace_id, file_no, payload, skipped, "NON-IR", dry_run)
        if result == "inserted":
            inserted += 1
        elif result == "updated":
            updated += 1
        else:
            skipped_count += 1

    print(f"[NON-IR]       {'Would insert' if dry_run else 'Inserted'}: {inserted} | {'Would update' if dry_run else 'Updated'}: {updated} | Skipped: {skipped_count}")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    args = sys.argv[1:]
    dry_run = "--dry-run" in args
    positional = [a for a in args if not a.startswith("--")]
    excel_path = positional[0] if positional else DEFAULT_EXCEL_PATH

    if not os.path.exists(excel_path):
        raise SystemExit(f"Excel file not found: {excel_path}")

    if dry_run:
        print("*** DRY RUN — no changes will be written to the database ***\n")
        print_mapping_report()

    print(f"Loading: {excel_path}")
    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)

    sb = create_client(SUPABASE_URL, SERVICE_ROLE_KEY)

    res = sb.table("votum_users").select("workspace_id").eq("email", WORKSPACE_OWNER_EMAIL).limit(1).execute()
    if not res.data:
        raise SystemExit(f"No user found for {WORKSPACE_OWNER_EMAIL}")
    workspace_id = res.data[0]["workspace_id"]
    print(f"Workspace: {workspace_id}\n")

    skipped = []

    process_pendency(wb["FInal Pendency"], sb, workspace_id, skipped, dry_run)
    process_scn(wb["SCN"], sb, workspace_id, skipped, dry_run)
    process_non_ir(wb["NON-IR"], sb, workspace_id, skipped, dry_run)

    if skipped:
        if not dry_run:
            with open(SKIPPED_CSV, "w", newline="") as f:
                writer = csv.DictWriter(f, fieldnames=list(skipped[0].keys()))
                writer.writeheader()
                writer.writerows(skipped)
            print(f"\nSkipped {len(skipped)} rows — see {SKIPPED_CSV}")
        else:
            print(f"\nWould skip {len(skipped)} rows (errors during DB lookup).")
    else:
        print("\nNo rows skipped.")

    print("\nDry run complete — no data written." if dry_run else "\nDone.")


if __name__ == "__main__":
    main()
