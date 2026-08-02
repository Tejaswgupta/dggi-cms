"""
Ingest IR cases from 'MZU_ Cases booked (IRs issued) 26-27.xlsx' into dggi_records.

Each sheet covers one financial year. Column layouts differ by era:

  NEW layout (2023-24 to 2026-27):
    col 0  Sr. No.
    col 1  335-J Register Number   → record_id (upsert key 1)
    col 2  Name of the Party       → taxpayer_name
    col 3  Address of the Party    → (skipped)
    col 4  GSTIN/PAN/Aadhar        → gstins
    col 5  Date of Detection       → date_of_ir + date_of_initiation
    col 6  Detection (Lakhs)       → detection_amount (× 1,00,000)
    col 7  Recovery (Lakhs)        → recovery_itc  (× 1,00,000)
    col 8-10 Seizure cols          → (skipped)
    col 11 Gist of case            → issue_involved
    -- 2023-24 ends here: col12=SIO, col13=Group
    -- 2024-25: col12-16 extra cols, col17=SIO, col18=Group, col19=Digit
    -- 2025-26/2026-27: col12-18 extra, col19=Mode, col20=SIO, col21=Group, col22=Digit

  OLD layout (2022-23 and earlier):
    col 0  Sr. No.
    col 1  335-J Register Number   → record_id (upsert key 1)
    col 2  Name and Address        → taxpayer_name (no separate GSTIN col)
    col 3  GSTIN/PAN               → gstins       (only for 2022-23 era)
    col 4  Date of Detection       → date_of_ir + date_of_initiation
    col 5  Detection (Lakhs)       → detection_amount (× 1,00,000)
    col 6  Recovery (Lakhs)        → recovery_itc  (× 1,00,000)
    col 11 Gist                    → issue_involved  (2022-23, 2021-22, 2020-21, 2019-20, 2017-18)
    col 7  Gist                    → issue_involved  (2018-19 only — no seizure sub-cols)
    col 11/12 SIO/Group            → sio_name, group (era-dependent, see SHEET_CONFIGS)

Dedup: match on record_id (335-J number) first; if it's an integer, format as
  {int}/{GST|ST}/{YY-YY}  (best-effort, skips if ambiguous).
  Fallback: digit_id match.
  Final fallback: insert.

Usage:
    python3 scripts/ingest_cases_booked.py [/path/to/file.xlsx] [--dry-run] [--sheets 2026-27,2025-26]
"""

import csv
import json
import os
import re
import sys
from datetime import date, datetime

import openpyxl
from dotenv import load_dotenv
from supabase import create_client

load_dotenv(os.path.join(os.path.dirname(__file__), ".env"))
SUPABASE_URL = os.environ["SUPABASE_URL"]
SERVICE_ROLE_KEY = os.environ["SERVICE_ROLE_KEY"]

WORKSPACE_OWNER_EMAIL = "ajinkya.k1@gov.in"

DEFAULT_EXCEL_PATH = os.path.join(
    os.path.dirname(__file__), "..", "data",
    "MZU_ Cases booked (IRs issued) 26-27.xlsx",
)

LOG_JSON    = os.path.join(os.path.dirname(__file__), "ingest_cases_booked_log.json")
SKIPPED_CSV = os.path.join(os.path.dirname(__file__), "ingest_cases_booked_skipped.csv")

# ── Sheet configs ─────────────────────────────────────────────────────────────
# Keys: (col_taxpayer, col_gstin, col_date, col_detection, col_recovery,
#        col_gist, col_sio, col_group, col_digit, col_mode, has_address_col)
# col_digit / col_mode = None when not present in that sheet.
# has_address_col: if True, col2=taxpayer, col3=address, col4=gstin (new layout)
#                 if False, col2=name+address, col3=gstin (old layout)

SHEET_CONFIGS = {
    # sheet_name:  (taxpayer, gstin, date, detect, recover, gist, sio, group, digit, mode)
    "IR Issued (2026-27)": dict(
        taxpayer=2, gstin=4, date=5, detection=6, recovery=7, gist=11,
        sio=20, group=21, digit=22, mode=19, has_address=True,
    ),
    "IR Issued (2025-26)": dict(
        taxpayer=2, gstin=4, date=5, detection=6, recovery=7, gist=11,
        sio=20, group=21, digit=22, mode=19, has_address=True,
    ),
    "IR Issued (2024-25)": dict(
        taxpayer=2, gstin=4, date=5, detection=6, recovery=7, gist=11,
        sio=17, group=18, digit=19, mode=None, has_address=True,
    ),
    "IR Issued (2023-24)": dict(
        taxpayer=2, gstin=4, date=5, detection=6, recovery=7, gist=11,
        sio=12, group=13, digit=None, mode=None, has_address=True,
    ),
    "IR Issued (2022-23)": dict(
        taxpayer=2, gstin=3, date=4, detection=5, recovery=6, gist=10,
        sio=11, group=12, digit=None, mode=None, has_address=False,
    ),
    "IR Issued (2021-22)": dict(
        taxpayer=2, gstin=3, date=4, detection=5, recovery=6, gist=10,
        sio=11, group=12, digit=None, mode=None, has_address=False,
    ),
    "IR Issued (2020-21)": dict(
        taxpayer=2, gstin=3, date=4, detection=5, recovery=6, gist=10,
        sio=11, group=12, digit=None, mode=None, has_address=False,
    ),
    "IR issued (2019-20)": dict(
        taxpayer=2, gstin=3, date=4, detection=5, recovery=6, gist=10,
        sio=None, group=11, digit=None, mode=None, has_address=False,
    ),
    "IR issued (2018-19)": dict(
        # No seizure sub-cols — gist is col7, group=col8, no SIO col
        taxpayer=2, gstin=3, date=4, detection=5, recovery=6, gist=7,
        sio=None, group=8, digit=None, mode=None, has_address=False,
    ),
    "IR issued (2017-18)": dict(
        taxpayer=2, gstin=3, date=4, detection=5, recovery=6, gist=10,
        sio=None, group=11, digit=None, mode=None, has_address=False,
    ),
}

# Mode mapping — raw cell values → DB enum values
MODE_MAP = {
    "1. summons": "Summons", "summon": "Summons", "summons": "Summons",
    "2. inspection": "Inspection", "inspection": "Inspection",
    "3. search": "Search", "search": "Search",
    "4. letter/email etc.": "Letter", "letter": "Letter", "email": "Email",
}

# ── Helpers ───────────────────────────────────────────────────────────────────

def parse_date(val) -> str | None:
    if val is None:
        return None
    if isinstance(val, (datetime, date)):
        d = val.date() if isinstance(val, datetime) else val
        if d > date.today() and d.day <= 12:
            try:
                d = d.replace(month=d.day, day=d.month)
            except ValueError:
                pass
        if d > date.today():
            return None
        return d.isoformat()
    s = str(val).strip()
    if not s:
        return None
    for fmt in ("%d.%m.%Y", "%d/%m/%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(s, fmt).date().isoformat()
        except ValueError:
            pass
    return None


def clean(val) -> str | None:
    if val is None:
        return None
    s = str(val).strip()
    return s if s else None


def str_amount(val) -> str | None:
    """Lakhs → absolute rupees string."""
    if val is None:
        return None
    s = str(val).strip()
    if s in ("-", "", "NA", "N/A"):
        return None
    try:
        f = float(s)
        return str(f * 100_000) if f != 0 else None
    except (ValueError, TypeError):
        return None


def normalize_ir_no(val, sheet_fy: str | None = None) -> str | None:
    """Normalize col1 to a string record_id.

    Integer values (e.g. 33) are treated as raw 335-J serial numbers and
    reconstructed as '{n}/GST/{fy}' when the sheet FY is known — this lets
    the upsert match them in the DB if they exist in that format.
    """
    if val is None:
        return None
    if isinstance(val, int):
        if sheet_fy:
            return f"{val}/GST/{sheet_fy}"
        return None
    s = str(val).strip()
    return s if s else None


def normalize_group(val) -> str | None:
    if val is None:
        return None
    s = str(val).strip()
    # Already "Group X" form
    if s.startswith("Group "):
        return s
    # Single letter
    if re.match(r"^[A-Fa-f]$", s):
        return f"Group {s.upper()}"
    return None


def normalize_mode(val) -> str | None:
    if val is None:
        return None
    s = str(val).strip().lower()
    return MODE_MAP.get(s) or next(
        (v for k, v in MODE_MAP.items() if k in s), None
    )


def _is_real_digit_id(val: str | None) -> bool:
    if not val:
        return False
    s = val.strip()
    if s.lower() in ("ok", "not in digit", "yes", "no", ""):
        return False
    return bool(re.match(r"^\d{10,}-\d+$", s)) or bool(re.match(r"^DIGIT-\d", s, re.IGNORECASE))


def strip_digit_prefix(digit_id: str | None) -> str | None:
    if not digit_id:
        return None
    if digit_id.upper().startswith("DIGIT-"):
        return digit_id[6:]
    return digit_id


# ── Upsert logic ──────────────────────────────────────────────────────────────

def insert_record(sb, workspace_id, sr_no, ir_no, digit_id_raw, payload,
                  sheet_name, skipped, log, dry_run):
    try:
        # 1. Skip if record_id already exists
        if ir_no:
            res = (sb.table("dggi_records")
                   .select("id,record_id")
                   .eq("workspace_id", workspace_id)
                   .eq("record_id", ir_no)
                   .execute())
            if res.data:
                row = res.data[0]
                log.append({"action": "skip", "reason": "record_id exists", "sr_no": sr_no,
                            "sheet": sheet_name, "record_id": row["record_id"], "db_id": row["id"]})
                if dry_run:
                    print(f"    → SKIP (exists by ir_no)  record_id={row['record_id']!r}")
                return "skipped"

        # 2. Skip if digit_id already exists
        digit_id_db = strip_digit_prefix(digit_id_raw)
        if digit_id_db:
            res = (sb.table("dggi_records")
                   .select("id,record_id")
                   .eq("workspace_id", workspace_id)
                   .eq("digit_id", digit_id_db)
                   .execute())
            if res.data:
                row = res.data[0]
                log.append({"action": "skip", "reason": "digit_id exists", "sr_no": sr_no,
                            "sheet": sheet_name, "record_id": row["record_id"], "db_id": row["id"],
                            "digit_id": digit_id_raw})
                if dry_run:
                    print(f"    → SKIP (exists by digit_id)  record_id={row['record_id']!r}")
                return "skipped"

        # 3. Insert new record
        new_id = ir_no or f"CB-{sheet_name[-7:].replace('-', '')}-{sr_no:03d}"
        insert_payload = {**payload, "workspace_id": workspace_id, "record_id": new_id}
        if dry_run:
            print(f"    → INSERT  record_id={new_id!r}")
            db_id = None
        else:
            res = sb.table("dggi_records").insert(insert_payload).execute()
            db_id = res.data[0]["id"] if res.data else None
        log.append({"action": "insert", "sr_no": sr_no, "sheet": sheet_name,
                    "record_id": new_id, "db_id": db_id, "digit_id": digit_id_raw})
        return "inserted"

    except Exception as e:
        skipped.append({"sr_no": sr_no, "sheet": sheet_name, "ir_no": ir_no or "",
                        "digit_id": digit_id_raw or "", "reason": str(e)})
        if dry_run:
            print(f"    → ERROR  {e}")
        return "error"


# ── Sheet processor ───────────────────────────────────────────────────────────

def process_sheet(ws, sheet_name, sb, workspace_id, skipped, log, dry_run):
    cfg = SHEET_CONFIGS[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    inserted = updated = skipped_count = 0

    # Extract FY string from sheet name, e.g. "IR Issued (2025-26)" → "2025-26"
    m = re.search(r"\((\d{4}-\d{2,4})\)", sheet_name)
    sheet_fy = m.group(1) if m else None

    # Data starts at row index 2; skip row 0 (title) and row 1 (header).
    # Some sheets have a sub-header row at index 2 — skip if col0 is None or non-int.
    data_start = 2
    if rows[2][0] is None:
        data_start = 3

    for row in rows[data_start:]:
        if row[0] is None:
            continue
        try:
            sr_no = int(row[0])
        except (ValueError, TypeError):
            continue

        taxpayer = clean(row[cfg["taxpayer"]])
        if not taxpayer:
            continue

        gstin_col = cfg["gstin"]
        gstin = clean(row[gstin_col]) if gstin_col is not None else None

        date_val = parse_date(row[cfg["date"]])
        ir_no = normalize_ir_no(row[1], sheet_fy)

        detection = str_amount(row[cfg["detection"]])
        recovery  = str_amount(row[cfg["recovery"]])
        gist      = clean(row[cfg["gist"]])

        sio_col   = cfg.get("sio")
        sio_name  = clean(row[sio_col]) if sio_col is not None else None

        group_col = cfg.get("group")
        group     = normalize_group(row[group_col]) if group_col is not None else None

        digit_col    = cfg.get("digit")
        digit_id_raw = clean(row[digit_col]) if digit_col is not None else None
        if not _is_real_digit_id(digit_id_raw):
            digit_id_raw = None
        digit_id_db = strip_digit_prefix(digit_id_raw)

        mode_col = cfg.get("mode")
        mode     = normalize_mode(row[mode_col]) if mode_col is not None else None

        payload = {
            "taxpayer_name":    taxpayer,
            "gstins":           gstin,
            "date_of_ir":       date_val,
            "date_of_initiation": date_val,
            "detection_amount": detection,
            "recovery_itc":     recovery,
            "issue_involved":   gist,
            "sio_name":         sio_name,
            "group":            group,
            "digit_id":         digit_id_db,
            "mode_of_initiation": mode,
            "is_ir":            True,
            "closure_by":       "Closure Report Filed",
        }
        payload = {k: v for k, v in payload.items() if v is not None}

        if dry_run:
            print(
                f"  [#{sr_no:03d}] {taxpayer[:40]!r}"
                f" | ir_no={ir_no!r}"
                f" | digit={digit_id_raw!r}"
                f" | group={group!r}"
            )

        result = insert_record(sb, workspace_id, sr_no, ir_no, digit_id_raw,
                               payload, sheet_name, skipped, log, dry_run)
        if result == "inserted":
            inserted += 1
        elif result == "skipped":
            skipped_count += 1
        else:
            skipped_count += 1

    tag = "Would " if dry_run else ""
    print(
        f"  [{sheet_name}]  "
        f"{tag}Insert: {inserted}"
        f"  Already exists (skipped): {skipped_count}"
    )
    return inserted, skipped_count


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    args = sys.argv[1:]
    dry_run = "--dry-run" in args
    positional = [a for a in args if not a.startswith("--")]
    sheets_arg = next((a.split("=", 1)[1] for a in args if a.startswith("--sheets=")), None)

    excel_path = os.path.abspath(positional[0] if positional else DEFAULT_EXCEL_PATH)
    if not os.path.exists(excel_path):
        raise SystemExit(f"Excel file not found: {excel_path}")

    if dry_run:
        print("*** DRY RUN — no changes will be written ***\n")

    print(f"Loading: {excel_path}")
    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)

    # Determine which sheets to process
    if sheets_arg:
        # --sheets=2026-27,2025-26  →  match sheet names containing those FY strings
        fy_filters = [s.strip() for s in sheets_arg.split(",")]
        target_sheets = [
            s for s in SHEET_CONFIGS
            if any(fy in s for fy in fy_filters) and s in wb.sheetnames
        ]
    else:
        target_sheets = [s for s in SHEET_CONFIGS if "2026-27" in s and s in wb.sheetnames]

    if not target_sheets:
        raise SystemExit("No matching sheets found.")

    sb = create_client(SUPABASE_URL, SERVICE_ROLE_KEY)
    res = (sb.table("votum_users")
           .select("workspace_id")
           .eq("email", WORKSPACE_OWNER_EMAIL)
           .limit(1)
           .execute())
    if not res.data:
        raise SystemExit(f"No user found for {WORKSPACE_OWNER_EMAIL}")
    workspace_id = res.data[0]["workspace_id"]
    print(f"Workspace: {workspace_id}\n")

    skipped, log = [], []
    total_ins = total_skp = 0

    for sheet_name in target_sheets:
        print(f"\n{'='*60}")
        print(f"Processing: {sheet_name}")
        ins, skp = process_sheet(
            wb[sheet_name], sheet_name, sb, workspace_id, skipped, log, dry_run
        )
        total_ins += ins; total_skp += skp

    print(f"\n{'='*60}")
    print(f"TOTAL  Inserted: {total_ins}  Already existed (skipped): {total_skp}")

    if log and not dry_run:
        with open(LOG_JSON, "w") as f:
            json.dump(log, f, indent=2)
        print(f"\nLog → {LOG_JSON}")

    if skipped and not dry_run:
        with open(SKIPPED_CSV, "w", newline="") as f:
            writer = csv.DictWriter(f, fieldnames=list(skipped[0].keys()))
            writer.writeheader()
            writer.writerows(skipped)
        print(f"Skipped → {SKIPPED_CSV}")

    print("\n(dry run — no data written)" if dry_run else "\nDone.")


if __name__ == "__main__":
    main()
