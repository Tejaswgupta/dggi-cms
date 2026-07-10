
"""
Generate a sample Excel template for bulk user upload.
Client fills this out; we run seed_dd_users.py (or a future importer) from it.

Usage:
    python3 scripts/generate_user_upload_template.py
Output:
    scripts/user_upload_template.xlsx
"""

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

OUTPUT = "scripts/user_upload_template.xlsx"

# ── column definitions ────────────────────────────────────────────────────────
COLUMNS = [
    {
        "header": "Name",
        "key": "name",
        "width": 26,
        "required": True,
        "note": "Full name of the user",
    },
    {
        "header": "Email",
        "key": "email",
        "width": 34,
        "required": True,
        "note": "Official email address (must be unique)",
    },
    {
        "header": "Password",
        "key": "password",
        "width": 20,
        "required": True,
        "note": "Temporary login password (min 8 chars)",
    },
    {
        "header": "DGGI Role",
        "key": "dggi_role",
        "width": 18,
        "required": True,
        "note": "SIO / DD / ADD / ADG",
        "dropdown": ["SIO", "DD", "ADD", "ADG"],
    },
    {
        "header": "Group",
        "key": "group",
        "width": 16,
        "required": False,
        "note": "Group A–F (DD/SIO users only; leave blank for others)",
        "dropdown": ["Group A", "Group B", "Group C", "Group D", "Group E", "Group F"],
    },
]

SAMPLE_ROWS = [
    ["Rajiv Sharma",    "rajiv.sharma@dggi.gov.in",    "Dggi@1234", "SIO",  "Group A"],
    ["Meera Patel",     "meera.patel@dggi.gov.in",     "Dggi@1234", "DD",   "Group B"],
    ["Ankit Verma",     "ankit.verma@dggi.gov.in",     "Dggi@1234", "ADD",  ""],
    ["Sunita Nair",     "sunita.nair@dggi.gov.in",     "Dggi@1234", "ADG",  ""],
    ["Pradeep Kumar",   "pradeep.kumar@dggi.gov.in",   "Dggi@1234", "DD",   ""],
]

# ── colour palette ────────────────────────────────────────────────────────────
HEADER_BG      = "1F4E79"   # dark navy
REQUIRED_BG    = "FFF2CC"   # light yellow — required cols
OPTIONAL_BG    = "E9F5FF"   # light blue  — optional cols
SAMPLE_BG      = "F2F2F2"   # light grey  — sample rows
ALT_SAMPLE_BG  = "FFFFFF"   # white       — alternating
NOTE_BG        = "FFFBEA"
BORDER_COLOR   = "BFBFBF"


def thin_border():
    s = Side(style="thin", color=BORDER_COLOR)
    return Border(left=s, right=s, top=s, bottom=s)


def build_instructions(wb):
    ws = wb.create_sheet("Instructions")
    ws.sheet_view.showGridLines = False

    title_font  = Font(name="Calibri", bold=True, size=16, color="1F4E79")
    h2_font     = Font(name="Calibri", bold=True, size=12, color="1F4E79")
    body_font   = Font(name="Calibri", size=11)
    code_font   = Font(name="Courier New", size=10, color="C00000")

    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 60

    row = 2
    ws.cell(row, 2, "Bulk User Upload Template").font = title_font
    ws.cell(row, 2).alignment = Alignment(vertical="center")
    row += 2

    sections = [
        ("How to fill this sheet", [
            ("1.", "Switch to the 'Users' tab."),
            ("2.", "Delete the sample rows (rows 2–6) and add your own data."),
            ("3.", "Yellow columns are mandatory. Blue columns are optional."),
            ("4.", "Use the drop-downs in 'DGGI Role' and 'Group' columns."),
            ("5.", "Save as-is (.xlsx) and hand it back to the technical team."),
        ]),
        ("Field Rules", [
            ("Name",        "Full display name — no special characters."),
            ("Email",       "Must be a valid, unique email. Will be the login ID."),
            ("Password",    "Temporary password (min 8 chars). User changes on first login."),
            ("DGGI Role",   "SIO · DD · ADD · ADG — pick from the drop-down."),
            ("Group",       "Group A–F. Only fill for DD/SIO users. Leave blank otherwise."),
        ]),
        ("Notes", [
            ("•", "Do NOT change column headers or the sheet name 'Users'."),
            ("•", "Duplicate emails will be skipped during import."),
            ("•", "Passwords are temporary — advise users to reset after first login."),
        ]),
    ]

    for title, items in sections:
        c = ws.cell(row, 2, title)
        c.font = h2_font
        c.fill = PatternFill("solid", fgColor="D6E4F0")
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)
        row += 1
        for label, text in items:
            ws.cell(row, 2, label).font = code_font
            ws.cell(row, 3, text).font = body_font
            row += 1
        row += 1


def build_users(wb):
    ws = wb.active
    ws.title = "Users"
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A2"

    # ── header row ────────────────────────────────────────────────────────────
    header_font = Font(name="Calibri", bold=True, size=12, color="FFFFFF")
    for col_idx, col in enumerate(COLUMNS, start=1):
        cell = ws.cell(1, col_idx, col["header"] + (" *" if col["required"] else ""))
        cell.font = header_font
        cell.fill = PatternFill("solid", fgColor=HEADER_BG)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border()
        ws.column_dimensions[get_column_letter(col_idx)].width = col["width"]
    ws.row_dimensions[1].height = 32

    # ── sub-header (field notes) ──────────────────────────────────────────────
    note_font = Font(name="Calibri", italic=True, size=9, color="595959")
    for col_idx, col in enumerate(COLUMNS, start=1):
        cell = ws.cell(2, col_idx, col["note"])
        cell.font = note_font
        cell.fill = PatternFill("solid", fgColor=NOTE_BG)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border()
    ws.row_dimensions[2].height = 28

    # ── sample rows ───────────────────────────────────────────────────────────
    sample_font = Font(name="Calibri", size=11, italic=True, color="595959")
    for r_idx, row_data in enumerate(SAMPLE_ROWS, start=3):
        bg = SAMPLE_BG if r_idx % 2 == 1 else ALT_SAMPLE_BG
        for c_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(r_idx, c_idx, value)
            cell.font = sample_font
            cell.fill = PatternFill("solid", fgColor=bg)
            cell.alignment = Alignment(vertical="center")
            cell.border = thin_border()
        ws.row_dimensions[r_idx].height = 20

    # ── blank data rows (8–52) with column shading ────────────────────────────
    data_font = Font(name="Calibri", size=11)
    for r_idx in range(8, 53):
        bg = SAMPLE_BG if r_idx % 2 == 0 else ALT_SAMPLE_BG
        for c_idx, col in enumerate(COLUMNS, start=1):
            col_bg = REQUIRED_BG if col["required"] else OPTIONAL_BG
            cell = ws.cell(r_idx, c_idx)
            cell.font = data_font
            cell.fill = PatternFill("solid", fgColor=col_bg)
            cell.alignment = Alignment(vertical="center")
            cell.border = thin_border()
        ws.row_dimensions[r_idx].height = 20

    # ── drop-down validations ─────────────────────────────────────────────────
    for c_idx, col in enumerate(COLUMNS, start=1):
        if "dropdown" not in col:
            continue
        formula = '"' + ",".join(col["dropdown"]) + '"'
        dv = DataValidation(
            type="list",
            formula1=formula,
            allow_blank=not col["required"],
            showErrorMessage=True,
            errorTitle="Invalid value",
            error=f"Please choose from: {', '.join(col['dropdown'])}",
        )
        col_letter = get_column_letter(c_idx)
        dv.sqref = f"{col_letter}3:{col_letter}200"
        ws.add_data_validation(dv)

    # ── legend row at the top ─────────────────────────────────────────────────
    # (rows are 1-indexed; insert a visual note above header via sheet title row)
    legend_text = (
        "* = Required field   |   Yellow = Required   |   Blue = Optional   "
        "|   Sample rows are in grey — delete before submitting"
    )
    ws.insert_rows(1)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(COLUMNS))
    legend_cell = ws.cell(1, 1, legend_text)
    legend_cell.font = Font(name="Calibri", bold=True, size=10, color="FFFFFF")
    legend_cell.fill = PatternFill("solid", fgColor="2E75B6")
    legend_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 22


def main():
    wb = openpyxl.Workbook()
    build_users(wb)
    build_instructions(wb)

    # Make Users the active sheet on open
    wb.active = wb["Users"]

    wb.save(OUTPUT)
    print(f"Template written → {OUTPUT}")


if __name__ == "__main__":
    main()
