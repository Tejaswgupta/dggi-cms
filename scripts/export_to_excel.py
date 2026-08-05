"""
Export IR and Non-IR records from dggi_records to Excel files.

Output:
  data/export_ir_cases.xlsx
  data/export_non_ir_cases.xlsx
"""

import os
from datetime import datetime

from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from supabase import create_client

load_dotenv(os.path.join(os.path.dirname(__file__), ".env"))
SUPABASE_URL = os.environ["SUPABASE_URL"]
SERVICE_ROLE_KEY = os.environ["SERVICE_ROLE_KEY"]

WORKSPACE_ID = "e27632d5-19dc-49e6-92ec-df9a86567b40"

OUT_DIR = os.path.join(os.path.dirname(__file__), "..", "data")

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
DATA_ALIGN = Alignment(vertical="top", wrap_text=True)

# ── Column definitions ────────────────────────────────────────────────────────

IR_COLS = [
    ("Sr. No.",           None),
    ("IR No.",            "record_id"),
    ("Taxpayer Name",     "taxpayer_name"),
    ("GSTIN/PAN",         "gstins"),
    ("DIGIT ID",          "digit_id"),
    ("Date of IR",        "date_of_ir"),
    ("Detection (₹)",     "detection_amount"),
    ("Recovery ITC (₹)",  "recovery_itc"),
    ("Issue Involved",    "issue_involved"),
    ("Latest Status",     "latest_status"),
    ("SIO Name",          "sio_name"),
    ("Group",             "group"),
    ("Mode",              "mode_of_initiation"),
    ("Closure By",        "closure_by"),
]

NON_IR_COLS = [
    ("Sr. No.",           None),
    ("Record ID",         "record_id"),
    ("Non-IR Reg. No.",   "legacy_non_ir_no"),
    ("Taxpayer Name",     "taxpayer_name"),
    ("GSTIN/PAN",         "gstins"),
    ("Initiation Date",   "date_of_non_ir"),
    ("SIO Name",          "sio_name"),
    ("Group",             "group"),
    ("Mode",              "mode_of_initiation"),
    ("Latest Status",     "latest_status"),
]


def fmt(val):
    if val is None:
        return ""
    if isinstance(val, str) and val.endswith(".0") and val[:-2].lstrip("-").isdigit():
        return str(int(float(val)))
    return val


def write_sheet(ws, cols, rows):
    ws.row_dimensions[1].height = 30

    for col_idx, (header, _) in enumerate(cols, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGN

    for row_idx, record in enumerate(rows, 2):
        for col_idx, (_, field) in enumerate(cols, 1):
            if field is None:
                val = row_idx - 1
            else:
                val = fmt(record.get(field))
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.alignment = DATA_ALIGN

    # Auto-width (capped)
    for col_idx, (header, field) in enumerate(cols, 1):
        max_len = len(header)
        for row_idx in range(2, len(rows) + 2):
            v = ws.cell(row=row_idx, column=col_idx).value
            if v:
                max_len = max(max_len, min(len(str(v)), 60))
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = max_len + 2


def fetch_all(sb, is_ir: bool):
    all_rows = []
    page = 0
    page_size = 1000
    while True:
        res = (
            sb.table("dggi_records")
            .select("*")
            .eq("workspace_id", WORKSPACE_ID)
            .eq("is_ir", is_ir)
            .range(page * page_size, (page + 1) * page_size - 1)
            .execute()
        )
        all_rows.extend(res.data or [])
        if len(res.data or []) < page_size:
            break
        page += 1
    return all_rows


def main():
    sb = create_client(SUPABASE_URL, SERVICE_ROLE_KEY)

    # ── IR export ─────────────────────────────────────────────────────────────
    ir_rows = fetch_all(sb, is_ir=True)
    ir_rows.sort(key=lambda r: (r.get("date_of_ir") or "", r.get("record_id") or ""))

    wb_ir = Workbook()
    ws_ir = wb_ir.active
    ws_ir.title = "IR Cases"
    write_sheet(ws_ir, IR_COLS, ir_rows)

    ir_path = os.path.join(OUT_DIR, "export_ir_cases.xlsx")
    wb_ir.save(ir_path)
    print(f"IR export:     {ir_path}  ({len(ir_rows)} rows)")

    # ── Non-IR export ─────────────────────────────────────────────────────────
    nir_rows = fetch_all(sb, is_ir=False)
    nir_rows.sort(key=lambda r: (r.get("date_of_non_ir") or "", r.get("record_id") or ""))

    wb_nir = Workbook()
    ws_nir = wb_nir.active
    ws_nir.title = "Non-IR Cases"
    write_sheet(ws_nir, NON_IR_COLS, nir_rows)

    nir_path = os.path.join(OUT_DIR, "export_non_ir_cases.xlsx")
    wb_nir.save(nir_path)
    print(f"Non-IR export: {nir_path}  ({len(nir_rows)} rows)")


if __name__ == "__main__":
    main()
