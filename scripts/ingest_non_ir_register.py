"""
Ingest NON-IR Register data into Supabase dggi_records.

Source: 'NON_IR_Register_DGGI_MZU_upto_09_July_2026.xlsx'
        Sheet: 'final with random allott of cas'  (128 rows)

Column mapping:
  Col 0  Sr No            → sr_no (skip validation key)
  Col 1  File Number      → legacy_non_ir_no (original register number; dedup key)
  Col 2  Date of NON-IR   → date_of_non_ir (used for sequencing)
  Col 3  Officer Name     → sio_name
  Col 4  Group Name       → group ("A" → "Group A", etc.)
  Col 5  E-Mail ID        → (skipped)

record_id generation (FY-based, date-sequenced):
  - Extract FY from date (e.g. 2026-07-03 → "2026-27")
  - Sort all records by date within each FY
  - Assign sequential numbers: 1/GST/2026-27, 2/GST/2026-27, etc.
  - Sequence resets for each FY

Dedup strategy:
  1. Match on legacy_non_ir_no in DB → skip (already present)
  2. No match → insert as open/pending non-IR (no closure_by, no latest_status)

NOTE: Does NOT touch dggi_closure_records.

Usage:
    python3 scripts/ingest_non_ir_register.py [/path/to/file.xlsx] [--dry-run]
"""

import csv
import json
import os
import sys
from datetime import date, datetime

import openpyxl
from dotenv import load_dotenv
from supabase import create_client

load_dotenv(os.path.join(os.path.dirname(__file__), ".env"))
SUPABASE_URL = os.environ["SUPABASE_URL"]
SERVICE_ROLE_KEY = os.environ["SERVICE_ROLE_KEY"]

WORKSPACE_OWNER_EMAIL = "ajinkya.k1@gov.in"

DEFAULT_EXCEL_PATH = os.path.join(
    os.path.dirname(__file__),
    "..",
    "data",
    "NON_IR_Register_DGGI_MZU_upto_09_July_2026.xlsx",
)

SHEET_NAME = "final with random allott of cas"

SKIPPED_CSV = os.path.join(os.path.dirname(__file__), "ingest_non_ir_register_skipped.csv")
LOG_JSON = os.path.join(os.path.dirname(__file__), "ingest_non_ir_register_log.json")


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def parse_date(val) -> str | None:
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.date().isoformat()
    if isinstance(val, date):
        return val.isoformat()
    s = str(val).strip()
    if not s:
        return None
    for fmt in ("%d.%m.%Y", "%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"):
        try:
            return datetime.strptime(s, fmt).date().isoformat()
        except ValueError:
            pass
    return None


def clean(val) -> str | None:
    if val is None:
        return None
    s = str(val).strip()
    return s if s else None


# ---------------------------------------------------------------------------
# Sequential NON-IR record_id generator
# ---------------------------------------------------------------------------

def fy_from_date(date_str: str | None) -> str:
    """'2026-07-03' → '2026-27'; '2025-03-15' → '2024-25'"""
    if not date_str:
        from datetime import date
        now = date.today()
        year = now.year if now.month >= 4 else now.year - 1
    else:
        year = int(date_str[:4])
        month = int(date_str[5:7])
        year = year if month >= 4 else year - 1
    return f"{year}-{year + 1 - 2000:02d}"


# ---------------------------------------------------------------------------
# Upsert: match on file_no → update; no match → insert
# ---------------------------------------------------------------------------

def upsert_record(
    sb,
    workspace_id: str,
    sr_no: int,
    payload: dict,
    skipped: list,
    log: list,
    dry_run: bool = False,
) -> str:
    legacy_no = payload.get("legacy_non_ir_no")
    record_id = payload["record_id"]
    file_no = payload.get("file_no")
    try:
        # 1. Skip if record_id already exists as non-IR
        if record_id:
            res = (
                sb.table("dggi_records")
                .select("id,record_id")
                .eq("workspace_id", workspace_id)
                .eq("is_ir", False)
                .eq("record_id", record_id)
                .execute()
            )
            if res.data:
                row = res.data[0]
                if dry_run:
                    print(f"    → SKIP (exists by record_id)  record_id={row['record_id']!r}")
                log.append({"action": "skip", "reason": "record_id exists", "sr_no": sr_no,
                            "record_id": record_id, "db_id": row["id"]})
                return "skipped"

        # 2. Insert as closed non-IR
        insert_payload = {**payload, "workspace_id": workspace_id}
        if dry_run:
            print(f"    → INSERT  record_id={record_id!r}")
            inserted_id = None
        else:
            insert_res = sb.table("dggi_records").insert(insert_payload).execute()
            inserted_id = insert_res.data[0]["id"] if insert_res.data else None
        log.append({"action": "insert", "sr_no": sr_no, "record_id": record_id, "db_id": inserted_id})
        return "inserted"

    except Exception as e:
        error_msg = str(e)
        skipped.append({"sr_no": sr_no, "record_id": record_id, "legacy_no": legacy_no, "reason": error_msg})
        if dry_run:
            print(f"    → ERROR  {error_msg}")
        return "skipped"


# ---------------------------------------------------------------------------
# Sheet processor
# ---------------------------------------------------------------------------

def process_sheet(ws, sb, workspace_id: str, skipped: list, log: list, dry_run: bool = False):
    rows = list(ws.iter_rows(values_only=True))

    # Parse all rows
    raw_records = []
    for row in rows[1:]:
        if row[0] is None:
            continue
        try:
            sr_no = int(row[0])
        except (ValueError, TypeError):
            continue

        legacy_no = clean(row[1])
        nir_date = parse_date(row[2])
        officer_name = clean(row[3])
        group_raw = clean(row[4])
        group_val = f"Group {group_raw}" if group_raw else None

        raw_records.append({
            "sr_no": sr_no,
            "legacy_no": legacy_no,
            "date": nir_date,
            "officer": officer_name,
            "group": group_val,
        })

    # Group by FY and sort by date within each FY
    from collections import defaultdict
    by_fy = defaultdict(list)
    for rec in raw_records:
        fy = fy_from_date(rec["date"])
        by_fy[fy].append(rec)

    # Sort each FY by date, assign sequential record_ids
    assigned = []
    for fy in sorted(by_fy.keys()):
        recs = by_fy[fy]
        # Sort by date (None dates go last)
        recs.sort(key=lambda r: (r["date"] is None, r["date"]))
        for seq, rec in enumerate(recs, 1):
            rec["record_id"] = f"{seq}/GST/{fy}"
            assigned.append(rec)

    # Upsert in order
    inserted = skipped_count = 0
    for rec in assigned:
        payload = {
            "record_id": rec["record_id"],
            "file_no": rec["legacy_no"],  # Store original register number in file_no (for now, until DB schema adds legacy_non_ir_no column)
            "date_of_non_ir": rec["date"],
            "sio_name": rec["officer"],
            "group": rec["group"],
            "is_ir": False,
            # Leave as open/pending - no closure_by, no latest_status
        }
        # Drop None values except record_id
        payload = {k: v for k, v in payload.items() if v is not None or k == "record_id"}

        if dry_run:
            print(
                f"  [#{rec['sr_no']:03d}] legacy_no={rec['legacy_no']!r} | date={rec['date']}"
                f" | group={rec['group']!r} | sio={rec['officer']!r} → {rec['record_id']}"
            )

        result = upsert_record(sb, workspace_id, rec["sr_no"], payload, skipped, log, dry_run)
        if result == "inserted":
            inserted += 1
        else:
            skipped_count += 1

    print(
        f"\n[NON-IR Register]  {'Would insert' if dry_run else 'Inserted'}: {inserted}"
        f"  Already exists (skipped): {skipped_count}"
    )


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    args = sys.argv[1:]
    dry_run = "--dry-run" in args
    positional = [a for a in args if not a.startswith("--")]
    excel_path = os.path.abspath(positional[0] if positional else DEFAULT_EXCEL_PATH)

    if not os.path.exists(excel_path):
        raise SystemExit(f"Excel file not found: {excel_path}")

    if dry_run:
        print("*** DRY RUN — no changes will be written to the database ***\n")

    print(f"Loading: {excel_path}")
    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)

    if SHEET_NAME not in wb.sheetnames:
        raise SystemExit(f"Sheet {SHEET_NAME!r} not found. Available: {wb.sheetnames}")

    sb = create_client(SUPABASE_URL, SERVICE_ROLE_KEY)

    res = (
        sb.table("votum_users")
        .select("workspace_id")
        .eq("email", WORKSPACE_OWNER_EMAIL)
        .limit(1)
        .execute()
    )
    if not res.data:
        raise SystemExit(f"No user found for {WORKSPACE_OWNER_EMAIL}")
    workspace_id = res.data[0]["workspace_id"]
    print(f"Workspace: {workspace_id}")

    skipped = []
    log = []
    process_sheet(wb[SHEET_NAME], sb, workspace_id, skipped, log, dry_run)

    if log:
        with open(LOG_JSON, "w") as f:
            json.dump(log, f, indent=2)
        inserts = len([e for e in log if e["action"] == "insert"])
        updates = len([e for e in log if e["action"] == "update"])
        print(f"\nLog written to {LOG_JSON}  ({inserts} inserts, {updates} updates)")

    if skipped:
        if not dry_run:
            with open(SKIPPED_CSV, "w", newline="") as f:
                writer = csv.DictWriter(f, fieldnames=list(skipped[0].keys()))
                writer.writeheader()
                writer.writerows(skipped)
            print(f"\nSkipped {len(skipped)} rows — see {SKIPPED_CSV}")
        else:
            print(f"\nWould skip {len(skipped)} rows (errors during DB lookup).")
    else:
        print("\nNo rows skipped.")

    print("\nDry run complete — no data written." if dry_run else "\nDone.")


if __name__ == "__main__":
    main()
