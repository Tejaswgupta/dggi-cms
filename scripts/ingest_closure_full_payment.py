"""
Ingest IR Full-Payment-of-Dues closure data into Supabase dggi_closure_records.

Source: 'structured_closure_register_full_payment_of_dues (1) (1).xlsx'
        Sheet: 'Structured_Register'  (51 data rows, header at row 3)

Column mapping (0-indexed):
  Col 0  Financial Year / Series      → FY context (not stored directly)
  Col 1  Sr. No.                      → sr_no
  Col 2  File Number                  → file_no  (dedup key 1)
  Col 3  Taxpayer / Entity            → taxpayer_name
  Col 4  Closure Report Number        → (ignored — internal ref)
  Col 5  Closure Date (as seen)       → due_date
  Col 6  Section / Closure Provision  → issue_involved
  Col 7  Incident Report No.          → lookup dggi_records.record_id → source_record_id
  Col 8  Payment Details (as seen)    → recovery_cash  (parsed total int, rupees)
  Col 9  Officer / SIO / Group        → parse → sio_name + group
  Col 10 Source Page                  → (ignored)
  Col 11 Verification Status          → (ignored)
  Col 12 Issue / Note                 → (ignored)

Derived fields:
  is_ir          = True
  closure_by     = "Closed After Payment of Tax"
  source_record_id = looked up from dggi_records.record_id via IR No. (col 7)
                     fallback: dggi_records.file_no via col 2

Dedup strategy (checked in order):
  1. file_no match in dggi_closure_records → update in place
  2. No match → insert with new CIR/NNN/YY-YY record_id

record_id format: CIR/{seq:03d}/{fy}  e.g. CIR/001/26-27
  Sequence continues from max existing CIR record in the workspace.

Usage:
    python3 scripts/ingest_closure_full_payment.py [/path/to/file.xlsx] [--dry-run]
"""

import csv
import json
import os
import re
import sys
from datetime import date, datetime

import openpyxl
from dotenv import load_dotenv
from supabase import create_client

load_dotenv(os.path.join(os.path.dirname(__file__), ".env"))
SUPABASE_URL = os.environ["SUPABASE_URL"]
SERVICE_ROLE_KEY = os.environ["SERVICE_ROLE_KEY"]

WORKSPACE_OWNER_EMAIL = "ajinkya.k1@gov.in"

DEFAULT_EXCEL_PATH = os.path.join(
    os.path.dirname(__file__),
    "..",
    "data",
    "structured_closure_register_full_payment_of_dues.xlsx",
)

SHEET_NAME = "Structured_Register"
HEADER_ROW_INDEX = 3  # 0-indexed; row 4 in the sheet is the actual header

CLOSURE_BY = "Closed After Payment of Tax"

SKIPPED_CSV = os.path.join(os.path.dirname(__file__), "ingest_closure_full_payment_skipped.csv")
LOG_JSON = os.path.join(os.path.dirname(__file__), "ingest_closure_full_payment_log.json")


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def parse_date(val) -> str | None:
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.date().isoformat()
    if isinstance(val, date):
        return val.isoformat()
    s = str(val).strip().rstrip("?").strip()
    if not s:
        return None
    for fmt in ("%d/%m/%Y", "%d/%m/%y", "%d.%m.%Y", "%d.%m.%y", "%Y-%m-%d", "%d-%m-%Y"):
        try:
            return datetime.strptime(s, fmt).date().isoformat()
        except ValueError:
            pass
    return None


def clean(val) -> str | None:
    if val is None:
        return None
    s = str(val).strip()
    return s if s else None


_TOKEN_RE = re.compile(
    r"([\d,]+(?:\.\d+)?)\s*(crore?s?|lakh?s?|cr\.?)?(?=\s|/|,|\+|-|\Z)",
    re.IGNORECASE,
)

def parse_amount_to_int(text: str | None) -> int | None:
    """
    Parse a free-text Indian rupee amount into an integer (whole rupees).
    Handles crore/lakh multipliers, comma-separated numbers, and multi-component
    strings like 'Rs. 4,63,56,728/- along with int Rs. 86,97,196/-'.
    Strips parenthetical labels (Tax), (Int.) before parsing.
    Returns None if no numeric value is found.
    """
    if not text:
        return None
    cleaned = re.sub(r"\([^)]*\)", " ", text)
    total = 0.0
    found = False
    for m in _TOKEN_RE.finditer(cleaned):
        num_str = m.group(1).replace(",", "")
        unit = (m.group(2) or "").lower().rstrip(".")
        try:
            num = float(num_str)
        except ValueError:
            continue
        if unit in ("crore", "crores", "cr"):
            num *= 1e7
        elif unit in ("lakh", "lakhs"):
            num *= 1e5
        total += num
        found = True
    return round(total) if found else None


def fy_from_date(date_str: str | None) -> str:
    """'2026-05-01' → '26-27'"""
    if not date_str:
        now = date.today()
        s = now.year if now.month >= 4 else now.year - 1
    else:
        year, month = int(date_str[:4]), int(date_str[5:7])
        s = year if month >= 4 else year - 1
    return f"{str(s)[2:]}-{str(s + 1)[2:]}"


def parse_sio_group(raw: str | None) -> tuple[str | None, str | None]:
    """
    Parse strings like:
      'Ajay Kumar/ Gr.B', 'Gr-C/ YM', 'Rakesh K Singh/ Rahul Ghanghas Gr.E',
      'Vikash V./ Nitish Kaloya Grp E', 'Kapil Dev Gupta; SIO; Gr.C'
    Returns (officer_name, group_str) e.g. ('Ajay Kumar', 'Group B')
    """
    if not raw:
        return None, None

    # Extract group letter from Gr.B, Gr-C, Grp E, Gr.F(A), Group-A etc.
    group_match = re.search(r"\bGr[p]?[.\-\s]?\(?([A-F])\)?", raw, re.IGNORECASE)
    if not group_match:
        group_match = re.search(r"\bGroup[.\-\s]?([A-F])\b", raw, re.IGNORECASE)
    group_str = f"Group {group_match.group(1).upper()}" if group_match else None

    # Officer name: text before first comma/semicolon/slash, then iteratively
    # strip trailing role/group suffixes: "- Gr.B", "- SIO", "(IO)", etc.
    name_part = re.split(r"[,;/]", raw)[0].strip()
    _SUFFIX_RE = re.compile(
        r"([-\s]+(SIO|IO|Gr[p]?[.\-\s]*[A-F](\([A-Z]\))?)|\s*\([^)]*\))\s*$",
        re.IGNORECASE,
    )
    while True:
        stripped = _SUFFIX_RE.sub("", name_part).strip(" ,./;-()")
        if stripped == name_part:
            break
        name_part = stripped
    officer_name = name_part if name_part and len(name_part) > 1 else None

    return officer_name, group_str


# ---------------------------------------------------------------------------
# Build lookup caches
# ---------------------------------------------------------------------------

def build_user_cache_by_name(sb, workspace_id: str) -> dict:
    """Return {name_lower: user_id} for all users in the workspace."""
    res = (
        sb.table("votum_users")
        .select("id,name")
        .eq("workspace_id", workspace_id)
        .execute()
    )
    cache = {}
    for u in res.data:
        if u.get("name"):
            cache[u["name"].strip().lower()] = u["id"]
    return cache


def build_ir_source_cache(sb, workspace_id: str) -> tuple[dict, dict]:
    """
    Return two caches from dggi_records (IR cases):
      record_id_cache: {record_id_lower: record_id}  — for col 7 (IR No.) lookup
      file_no_cache:   {file_no_lower: record_id}    — for col 2 (file_no) fallback
    """
    res = (
        sb.table("dggi_records")
        .select("record_id,file_no")
        .eq("workspace_id", workspace_id)
        .eq("is_ir", True)
        .execute()
    )
    rid_cache = {}
    fn_cache = {}
    for r in res.data:
        if r.get("record_id"):
            rid_cache[r["record_id"].strip().lower()] = r["record_id"]
        if r.get("file_no"):
            fn_cache[r["file_no"].strip().lower()] = r["record_id"]
    return rid_cache, fn_cache


# ---------------------------------------------------------------------------
# Sequential record_id generator
# ---------------------------------------------------------------------------

def next_closure_ir_seq(sb, workspace_id: str) -> int:
    res = (
        sb.table("dggi_closure_records")
        .select("record_id")
        .eq("workspace_id", workspace_id)
        .eq("is_ir", True)
        .like("record_id", "CIR/%")
        .execute()
    )
    max_seq = 0
    for r in res.data:
        m = re.match(r"CIR/(\d+)/", r["record_id"] or "")
        if m:
            max_seq = max(max_seq, int(m.group(1)))
    return max_seq + 1


# ---------------------------------------------------------------------------
# Upsert logic
# ---------------------------------------------------------------------------

def upsert_closure_record(
    sb,
    workspace_id: str,
    sr_no: str,
    file_no: str | None,
    payload: dict,
    skipped: list,
    log: list,
    dry_run: bool = False,
) -> str:
    new_record_id = payload["record_id"]
    try:
        # 1. Match on file_no in dggi_closure_records → update
        if file_no:
            res = (
                sb.table("dggi_closure_records")
                .select("id,record_id")
                .eq("workspace_id", workspace_id)
                .eq("file_no", file_no)
                .execute()
            )
            if res.data:
                row = res.data[0]
                if dry_run:
                    print(f"    → UPDATE (file_no)  existing record_id={row['record_id']!r} → {new_record_id!r}  db_id={row['id']}")
                else:
                    sb.table("dggi_closure_records").update(payload).eq("id", row["id"]).execute()
                log.append({
                    "action": "update",
                    "match_by": "file_no",
                    "sr_no": sr_no,
                    "file_no": file_no,
                    "old_record_id": row["record_id"],
                    "new_record_id": new_record_id,
                    "db_id": row["id"],
                })
                return "updated"

        # 2. Insert
        insert_payload = {**payload, "workspace_id": workspace_id}
        if dry_run:
            print(f"    → INSERT  new record_id={new_record_id!r}")
            inserted_id = None
        else:
            insert_res = sb.table("dggi_closure_records").insert(insert_payload).execute()
            inserted_id = insert_res.data[0]["id"] if insert_res.data else None
        log.append({
            "action": "insert",
            "sr_no": sr_no,
            "new_record_id": new_record_id,
            "db_id": inserted_id,
            "file_no": file_no,
        })
        return "inserted"

    except Exception as e:
        skipped.append({"sr_no": sr_no, "file_no": file_no or "", "reason": str(e)})
        return "skipped"


# ---------------------------------------------------------------------------
# Sheet processor
# ---------------------------------------------------------------------------

def process_sheet(
    ws,
    sb,
    workspace_id: str,
    user_cache: dict,
    rid_cache: dict,
    fn_cache: dict,
    start_seq: int,
    skipped: list,
    log: list,
    dry_run: bool = False,
):
    rows = list(ws.iter_rows(values_only=True))

    # Collect valid data rows (after header row at HEADER_ROW_INDEX)
    records = []
    for row in rows[HEADER_ROW_INDEX + 1:]:
        if all(v is None for v in row):
            continue
        file_no = clean(row[2]) if len(row) > 2 else None
        taxpayer = clean(row[3]) if len(row) > 3 else None
        if not file_no and not taxpayer:
            continue

        sr_no = clean(row[1]) or "?"
        closure_date = parse_date(row[5]) if len(row) > 5 else None
        ir_no_raw = clean(row[7]) if len(row) > 7 else None
        sio_group_raw = clean(row[9]) if len(row) > 9 else None
        officer_name, group_str = parse_sio_group(sio_group_raw)

        records.append({
            "sr_no": sr_no,
            "file_no": file_no,
            "taxpayer_name": taxpayer,
            "closure_date": closure_date,
            "section": clean(row[6]) if len(row) > 6 else None,
            "ir_no_raw": ir_no_raw,
            "payment_raw": clean(row[8]) if len(row) > 8 else None,
            "officer_name": officer_name,
            "group": group_str,
        })

    # Sort by closure date ascending; no-date rows go last
    records.sort(key=lambda r: r["closure_date"] or "9999-99-99")

    inserted = updated = skipped_count = 0
    unmatched_officers = set()
    unmatched_ir_nos = set()

    for idx, rec in enumerate(records):
        # Per-record FY from its closure date
        fy = fy_from_date(rec["closure_date"])
        new_record_id = f"CIR/{idx + start_seq:03d}/{fy}"
        file_no = rec["file_no"]
        ir_no_raw = rec["ir_no_raw"]

        # Resolve officer UUID by name
        handling_io_sio_id = None
        if rec["officer_name"]:
            handling_io_sio_id = user_cache.get(rec["officer_name"].lower())
            if not handling_io_sio_id:
                unmatched_officers.add(rec["officer_name"])

        # Resolve source_record_id:
        # Primary — IR No. (col 7) matched against dggi_records.record_id
        # Fallback — file_no matched against dggi_records.file_no
        source_record_id = None
        if ir_no_raw:
            source_record_id = rid_cache.get(ir_no_raw.strip().lower())
        if not source_record_id and file_no:
            source_record_id = fn_cache.get(file_no.strip().lower())
        if ir_no_raw and not source_record_id:
            unmatched_ir_nos.add(ir_no_raw)

        payload = {
            "record_id": new_record_id,
            "source_record_id": source_record_id,
            "is_ir": True,
            "file_no": file_no,
            "taxpayer_name": rec["taxpayer_name"],
            "due_date": rec["closure_date"],
            "issue_involved": rec["section"],
            "recovery_cash": parse_amount_to_int(rec["payment_raw"]),
            "group": rec["group"],
            "handling_io_sio": handling_io_sio_id,
            "closure_by": CLOSURE_BY,
        }
        ALWAYS_INCLUDE = {"record_id", "is_ir", "closure_by"}
        payload = {k: v for k, v in payload.items() if v is not None or k in ALWAYS_INCLUDE}

        if dry_run:
            src = source_record_id or "NO MATCH"
            ir_str = f"ir_no={ir_no_raw!r} → {src!r}"
            officer_str = f"{rec['officer_name']!r} → {handling_io_sio_id or 'NO MATCH'}"
            amt = parse_amount_to_int(rec["payment_raw"])
            print(
                f"  [#{rec['sr_no']}] file_no={file_no!r}\n"
                f"          taxpayer={rec['taxpayer_name']!r}\n"
                f"          date={rec['closure_date']} | group={rec['group']!r}\n"
                f"          officer={officer_str}\n"
                f"          payment={rec['payment_raw']!r} → {amt}\n"
                f"          {ir_str} → {new_record_id}"
            )

        result = upsert_closure_record(
            sb, workspace_id, rec["sr_no"], file_no, payload, skipped, log, dry_run
        )
        if result == "inserted":
            inserted += 1
        elif result == "updated":
            updated += 1
        else:
            skipped_count += 1

    print(
        f"\n[Closure Full Payment]  "
        f"{'Would insert' if dry_run else 'Inserted'}: {inserted}"
        f" | {'Would update' if dry_run else 'Updated'}: {updated}"
        f" | Skipped: {skipped_count}"
    )

    if unmatched_officers:
        print(f"\nWarning: {len(unmatched_officers)} officer name(s) not found in votum_users (handling_io_sio left null):")
        for name in sorted(unmatched_officers):
            print(f"  {name!r}")

    if unmatched_ir_nos:
        print(f"\nNote: {len(unmatched_ir_nos)} IR No(s) not found in dggi_records (source_record_id left null or resolved via file_no):")
        for ir in sorted(unmatched_ir_nos):
            print(f"  {ir!r}")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    args = sys.argv[1:]
    dry_run = "--dry-run" in args
    positional = [a for a in args if not a.startswith("--")]
    excel_path = os.path.abspath(positional[0] if positional else DEFAULT_EXCEL_PATH)

    if not os.path.exists(excel_path):
        raise SystemExit(f"Excel file not found: {excel_path}")

    if dry_run:
        print("*** DRY RUN — no changes will be written to the database ***\n")

    print(f"Loading: {excel_path}")
    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)

    if SHEET_NAME not in wb.sheetnames:
        raise SystemExit(f"Sheet {SHEET_NAME!r} not found. Available: {wb.sheetnames}")

    sb = create_client(SUPABASE_URL, SERVICE_ROLE_KEY)

    res = (
        sb.table("votum_users")
        .select("workspace_id")
        .eq("email", WORKSPACE_OWNER_EMAIL)
        .limit(1)
        .execute()
    )
    if not res.data:
        raise SystemExit(f"No user found for {WORKSPACE_OWNER_EMAIL}")
    workspace_id = res.data[0]["workspace_id"]
    print(f"Workspace: {workspace_id}")

    user_cache = build_user_cache_by_name(sb, workspace_id)
    print(f"Loaded {len(user_cache)} users from votum_users (by name)")

    rid_cache, fn_cache = build_ir_source_cache(sb, workspace_id)
    print(f"Loaded {len(rid_cache)} IR records by record_id, {len(fn_cache)} by file_no from dggi_records")

    start_seq = next_closure_ir_seq(sb, workspace_id)
    print(f"Next CIR sequence starts at: {start_seq}\n")

    skipped = []
    log = []
    process_sheet(
        wb[SHEET_NAME], sb, workspace_id,
        user_cache, rid_cache, fn_cache, start_seq,
        skipped, log, dry_run
    )

    if log:
        with open(LOG_JSON, "w") as f:
            json.dump(log, f, indent=2)
        inserts = len([e for e in log if e["action"] == "insert"])
        updates = len([e for e in log if e["action"] == "update"])
        print(f"\nLog written to {LOG_JSON}  ({inserts} inserts, {updates} updates)")

    if skipped:
        if not dry_run:
            with open(SKIPPED_CSV, "w", newline="") as f:
                writer = csv.DictWriter(f, fieldnames=list(skipped[0].keys()))
                writer.writeheader()
                writer.writerows(skipped)
            print(f"\nSkipped {len(skipped)} rows — see {SKIPPED_CSV}")
        else:
            print(f"\nWould skip {len(skipped)} rows (errors during DB lookup).")
    else:
        print("\nNo rows skipped.")

    print("\nDry run complete — no data written." if dry_run else "\nDone.")


if __name__ == "__main__":
    main()
