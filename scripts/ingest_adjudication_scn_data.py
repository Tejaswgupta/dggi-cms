"""
Ingest Adjudication Status SCN Register data into Supabase dggi_scn_records.

Source: 'Adjudication Status SCN Register Entry All Grp.xlsx'
Sheets: 'DD & AD', 'SIO', 'ADD & JD '  (all have identical column layout)

Row layout per sheet:
  Row 0  Header labels
  Row 1  Sub-headers (Tax / Interest / Penalty under Demand)
  Row 2  Column numbers — data begins at row 3

Column mapping:
  Col 0  SR NO.                                              → upsert sequence / skip-check
  Col 1  SCN No.                                            → record_id  (upsert key)
  Col 2  Date of SCN                                        → date_of_scn
  Col 3  Name of each Noticee                               → taxpayer_name
  Col 4  GSTIN/PAN                                          → gstins
  Col 5  Demand – Tax (in Rs.)                              → demand_tax
  Col 6  Demand – Interest (in Rs.)                         → demand_interest
  Col 7  Demand – Penalty (in Rs.)                          → demand_penalty
  Col 8  Period Involved                                    → period_involved
  Col 9  Last date of OIO                                   → last_date_of_oio
  Col 10 Issue                                              → issue_involved
  Col 11 Adjudication Formation                             → adjudication_formation
  Col 12 File No.                                           → file_no
  Col 13 DIN No.                                            → din_no
  Col 14 Date of uploading on BO Portal                    → date_of_bo_portal_upload
  Col 15 Adjudication Status as per BO Portal              → adjudication_status
  Col 16 Remarks, if any                                    → remarks
  Col 17 Date                                               → date_of_adjudication
  Col 18 Group Name                                         → group_name
  Col 19 Officer Name                                       → officer_name
  Sheet name                                                → source_sheet

Dedup strategy (checked in order):
  1. SCN No. match — exact match of col 1 vs record_id in DB  (primary)
  2. No match → insert; record_id = SCN No. (fallback: "{sheet_abbr}-{sr_no:03d}")

Usage:
    python3 scripts/ingest_adjudication_scn_data.py [/path/to/file.xlsx] [--dry-run]
"""

import csv
import json
import os
import sys
from datetime import date, datetime

import openpyxl
from dotenv import load_dotenv
from supabase import create_client

load_dotenv(os.path.join(os.path.dirname(__file__), ".env"))
SUPABASE_URL = os.environ["SUPABASE_URL"]
SERVICE_ROLE_KEY = os.environ["SERVICE_ROLE_KEY"]

WORKSPACE_OWNER_EMAIL = "ajinkya.k1@gov.in"

DEFAULT_EXCEL_PATH = os.path.join(
    os.path.dirname(__file__),
    "..",
    "data",
    "Adjudication Status SCN Register Entry All Grp.xlsx",
)

SKIPPED_CSV = os.path.join(os.path.dirname(__file__), "ingest_adjudication_scn_skipped.csv")
LOG_JSON = os.path.join(os.path.dirname(__file__), "ingest_adjudication_scn_log.json")

SHEET_NAMES = ["DD & AD", "SIO", "ADD & JD "]

SHEET_COMPETENCY = {
    "DD & AD": "DD/AD",
    "SIO": "SIO",
    "ADD & JD ": "ADD/JD",
}


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def parse_date(val) -> str | None:
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.date().isoformat()
    if isinstance(val, date):
        return val.isoformat()
    s = str(val).strip()
    if not s or s in ("0", "-"):
        return None
    for fmt in ("%d.%m.%Y", "%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%d.%m.%y", "%d/%m/%y"):
        try:
            return datetime.strptime(s, fmt).date().isoformat()
        except ValueError:
            pass
    return None


def clean(val) -> str | None:
    if val is None:
        return None
    s = str(val).strip()
    if s in ("0", "-", "None"):
        return None
    return s if s else None


def parse_amount(val) -> str | None:
    """Parse demand amounts — handles both numeric and messy string formats like '65,39,84,609/-'."""
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return str(val) if val != 0 else None
    s = str(val).strip()
    if not s or s in ("-", "0", "None"):
        return None
    # Strip trailing "/- " characters and commas
    cleaned = s.replace(",", "").rstrip("/-").strip()
    try:
        f = float(cleaned)
        return str(f) if f != 0 else None
    except ValueError:
        return s if s else None


# ---------------------------------------------------------------------------
# Upsert: SCN No. → insert
# ---------------------------------------------------------------------------

def upsert_scn_record(
    sb,
    workspace_id: str,
    sr_no: int,
    scn_no: str | None,
    sheet_abbr: str,
    payload: dict,
    skipped: list,
    log: list,
    dry_run: bool = False,
) -> str:
    try:
        # 1. Match on SCN No. vs record_id
        if scn_no:
            res = (
                sb.table("dggi_scn_records")
                .select("id,record_id")
                .eq("workspace_id", workspace_id)
                .eq("record_id", scn_no)
                .execute()
            )
            if res.data:
                row = res.data[0]
                if dry_run:
                    print(f"    → UPDATE (scn_no)  existing record_id={row['record_id']!r}  db_id={row['id']}")
                else:
                    sb.table("dggi_scn_records").update(payload).eq("id", row["id"]).execute()
                log.append({"action": "update", "match_by": "scn_no", "sr_no": sr_no, "record_id": row["record_id"], "db_id": row["id"], "noticee_name": payload.get("noticee_name")})
                return "updated"

        # 2. Insert — use scn_no as record_id (fallback to sheet_abbr-{sr_no:03d})
        new_record_id = scn_no or f"{sheet_abbr}-{sr_no:03d}"
        insert_payload = {**payload, "workspace_id": workspace_id, "record_id": new_record_id}
        if dry_run:
            print(f"    → INSERT  new record_id={new_record_id!r}")
            inserted_id = None
        else:
            insert_res = sb.table("dggi_scn_records").insert(insert_payload).execute()
            inserted_id = insert_res.data[0]["id"] if insert_res.data else None
        log.append({"action": "insert", "sr_no": sr_no, "record_id": new_record_id, "db_id": inserted_id, "noticee_name": payload.get("noticee_name")})
        return "inserted"

    except Exception as e:
        skipped.append({"sr_no": sr_no, "scn_no": scn_no or "", "sheet": sheet_abbr, "reason": str(e)})
        return "skipped"


# ---------------------------------------------------------------------------
# Sheet processor
# ---------------------------------------------------------------------------

def process_sheet(ws, sheet_name: str, competency: str, sb, workspace_id: str, skipped: list, log: list, dry_run: bool = False):
    rows = list(ws.iter_rows(values_only=True))
    inserted = updated = skipped_count = 0

    # Rows 0–2 are header / sub-header / column-number rows; data starts at index 3
    for row in rows[3:]:
        if row[0] is None:
            continue
        try:
            sr_no = int(row[0])
        except (ValueError, TypeError):
            continue

        taxpayer_name = clean(row[3])  # used only for skip-check and dry-run label
        if not taxpayer_name:
            continue

        scn_no = clean(row[1])
        sheet_abbr = sheet_name.strip().replace(" ", "_").replace("&", "").replace("__", "_")

        payload = {
            "scn_no": scn_no,
            "date_of_scn": parse_date(row[2]),
            "noticee_name": taxpayer_name,
            "gstin_pan": clean(row[4]),
            "demand_tax": parse_amount(row[5]),
            "demand_interest": parse_amount(row[6]),
            "demand_penalty": parse_amount(row[7]),
            "period_involved": clean(row[8]),
            "last_date_oio": parse_date(row[9]),
            "issue": clean(row[10]),
            "adjudication_formation": clean(row[11]),
            "file_no": clean(row[12]),
            "din_no": clean(row[13]),
            "date_uploading_bo": parse_date(row[14]),
            "adjudication_status": clean(row[15]),
            "remarks": clean(row[16]),
            "competency": competency,
        }
        payload = {k: v for k, v in payload.items() if v is not None}

        if dry_run:
            print(
                f"  [#{sr_no:03d}] {taxpayer_name[:45]!r}"
                f" | scn_no={scn_no} | sheet={sheet_name.strip()!r}"
            )

        result = upsert_scn_record(
            sb, workspace_id, sr_no, scn_no, sheet_abbr, payload, skipped, log, dry_run
        )
        if result == "inserted":
            inserted += 1
        elif result == "updated":
            updated += 1
        else:
            skipped_count += 1

    print(
        f"\n[{sheet_name.strip()}]  {'Would insert' if dry_run else 'Inserted'}: {inserted}"
        f" | {'Would update' if dry_run else 'Updated'}: {updated}"
        f" | Skipped: {skipped_count}"
    )


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    args = sys.argv[1:]
    dry_run = "--dry-run" in args
    positional = [a for a in args if not a.startswith("--")]
    excel_path = os.path.abspath(positional[0] if positional else DEFAULT_EXCEL_PATH)

    if not os.path.exists(excel_path):
        raise SystemExit(f"Excel file not found: {excel_path}")

    if dry_run:
        print("*** DRY RUN — no changes will be written to the database ***\n")

    print(f"Loading: {excel_path}")
    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)

    sb = create_client(SUPABASE_URL, SERVICE_ROLE_KEY)

    res = (
        sb.table("votum_users")
        .select("workspace_id")
        .eq("email", WORKSPACE_OWNER_EMAIL)
        .limit(1)
        .execute()
    )
    if not res.data:
        raise SystemExit(f"No user found for {WORKSPACE_OWNER_EMAIL}")
    workspace_id = res.data[0]["workspace_id"]
    print(f"Workspace: {workspace_id}\n")

    if dry_run:
        print("*** DRY RUN — skipping truncate ***\n")
    else:
        print("Truncating existing dggi_scn_records for this workspace...")
        sb.table("dggi_scn_records").delete().eq("workspace_id", workspace_id).execute()
        print("Truncated.\n")

    skipped = []
    log = []

    for sheet_name in SHEET_NAMES:
        if sheet_name not in wb.sheetnames:
            print(f"[WARN] Sheet {sheet_name!r} not found — skipping.")
            continue
        competency = SHEET_COMPETENCY[sheet_name]
        process_sheet(wb[sheet_name], sheet_name, competency, sb, workspace_id, skipped, log, dry_run)

    if log:
        with open(LOG_JSON, "w") as f:
            json.dump(log, f, indent=2)
        inserts = len([e for e in log if e["action"] == "insert"])
        updates = len([e for e in log if e["action"] == "update"])
        print(f"\nLog written to {LOG_JSON}  ({inserts} inserts, {updates} updates)")

    if skipped:
        if not dry_run:
            with open(SKIPPED_CSV, "w", newline="") as f:
                writer = csv.DictWriter(f, fieldnames=list(skipped[0].keys()))
                writer.writeheader()
                writer.writerows(skipped)
            print(f"\nSkipped {len(skipped)} rows — see {SKIPPED_CSV}")
        else:
            print(f"\nWould skip {len(skipped)} rows (errors during DB lookup).")
    else:
        print("\nNo rows skipped.")

    print("\nDry run complete — no data written." if dry_run else "\nDone.")


if __name__ == "__main__":
    main()
