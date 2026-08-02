"""
Ingest Closure Reports data into Supabase dggi_closure_records.

Source: 'Closure Reports.xlsx'
  Sheet 'FP'  — Full Payment closures  (header row 1, data from row 2)
  Sheet 'NSP' — Non-Standard-Payment closures

FP column mapping (0-indexed):
  Col 0  Sr. No.                       → sr_no (skip)
  Col 1  File Number                   → file_no            (dedup key 2)
  Col 2  Taxpayer / Entity             → taxpayer_name
  Col 3  Closure Report Number         → record_id          (dedup key 1, use as-is)
  Col 4  Closure Date (as seen)        → due_date
  Col 5  Section / Closure Provision   → issue_involved
  Col 6  Incident Report No.           → lookup dggi_records.record_id → source_record_id
  Col 7  Payment Details (as seen)     → total_recovery
  Col 8  Group                         → group  (single letter → "Group X")
  Col 9  SIO                           → lookup votum_users.name → handling_io_sio

  Derived: is_ir = True, closure_by = "Closed After Payment of Tax"

NSP column mapping (0-indexed):
  Col 0  Sr. No.                       → sr_no (skip)
  Col 1  File Number                   → file_no            (dedup key 2)
  Col 2  Taxpayer / Entity             → taxpayer_name
  Col 3  GSTIN / Other ID              → gstins
  Col 4  Closure Report Number         → record_id          (dedup key 1, use as-is)
  Col 5  Closure Date (as seen)        → due_date
  Col 6  Section / U/S                 → issue_involved
  Col 7  Remark / Detection Status     → closure_reason
  Col 8  Group                         → group
  Col 9  SIO                           → lookup votum_users.name → handling_io_sio

  Derived: is_ir = True, closure_by = "Closed"

Dedup strategy (per row, in order):
  1. record_id exact match in dggi_closure_records → update
  2. file_no match in dggi_closure_records         → update
  3. No match                                      → insert

Usage:
    python3 scripts/ingest_closure_reports.py [/path/to/file.xlsx] [--dry-run]
"""

import csv
import json
import os
import re
import sys
from datetime import date, datetime

import openpyxl
from dotenv import load_dotenv
from supabase import create_client

load_dotenv(os.path.join(os.path.dirname(__file__), ".env"))
SUPABASE_URL = os.environ["SUPABASE_URL"]
SERVICE_ROLE_KEY = os.environ["SERVICE_ROLE_KEY"]

WORKSPACE_OWNER_EMAIL = "ajinkya.k1@gov.in"

DEFAULT_EXCEL_PATH = os.path.join(
    os.path.dirname(__file__),
    "..",
    "..",
    "Closure Reports.xlsx",
)

LOG_JSON = os.path.join(os.path.dirname(__file__), "ingest_closure_reports_log.json")
SKIPPED_CSV = os.path.join(os.path.dirname(__file__), "ingest_closure_reports_skipped.csv")


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def parse_date(val) -> str | None:
    if val is None:
        return None
    if isinstance(val, (datetime, date)):
        d = val.date() if isinstance(val, datetime) else val
        if d > date.today() and d.day <= 12:
            try:
                d = d.replace(month=d.day, day=d.month)
            except ValueError:
                pass
        if d > date.today():
            return None
        return d.isoformat()
    s = str(val).strip().rstrip("?").strip()
    if not s:
        return None
    # handle short year formats like "18/05/26"
    for fmt in ("%d/%m/%Y", "%d/%m/%y", "%d.%m.%Y", "%d.%m.%y", "%d-%m-%Y"):
        try:
            return datetime.strptime(s, fmt).date().isoformat()
        except ValueError:
            pass
    return None


def clean(val) -> str | None:
    if val is None:
        return None
    s = str(val).strip()
    return s if s else None


def normalize_group(raw: str | None) -> str | None:
    """'B' → 'Group B', 'Group B' → 'Group B', 'Gr.B' → 'Group B'"""
    if not raw:
        return None
    s = raw.strip()
    m = re.search(r"\b([A-F])\b", s, re.IGNORECASE)
    if m:
        return f"Group {m.group(1).upper()}"
    return None


_TOKEN_RE = re.compile(
    r"([\d,]+(?:\.\d+)?)\s*(crore?s?|lakh?s?|cr\.?)?(?=\s|/|,|\+|-|\Z)",
    re.IGNORECASE,
)


def parse_amount_to_int(text: str | None) -> int | None:
    """Parse free-text Indian rupee amounts into whole rupees integer."""
    if not text:
        return None
    cleaned = re.sub(r"\([^)]*\)", " ", text)
    total = 0.0
    found = False
    for m in _TOKEN_RE.finditer(cleaned):
        num_str = m.group(1).replace(",", "")
        unit = (m.group(2) or "").lower().rstrip(".")
        try:
            num = float(num_str)
        except ValueError:
            continue
        if unit in ("crore", "crores", "cr"):
            num *= 1e7
        elif unit in ("lakh", "lakhs"):
            num *= 1e5
        total += num
        found = True
    return round(total) if found else None


# ---------------------------------------------------------------------------
# Build lookup caches
# ---------------------------------------------------------------------------

def build_user_cache(sb, workspace_id: str) -> dict:
    """Return {name_lower: user_id} for all users in the workspace."""
    res = (
        sb.table("votum_users")
        .select("id,name")
        .eq("workspace_id", workspace_id)
        .execute()
    )
    cache = {}
    for u in res.data:
        if u.get("name"):
            cache[u["name"].strip().lower()] = u["id"]
    return cache


def build_source_cache(sb, workspace_id: str) -> tuple[dict, dict]:
    """
    Return two lookup caches from dggi_records:
      record_id_cache: {record_id_lower: record_id}
      file_no_cache:   {file_no_lower: record_id}
    """
    res = (
        sb.table("dggi_records")
        .select("record_id,file_no")
        .eq("workspace_id", workspace_id)
        .execute()
    )
    rid_cache: dict = {}
    fn_cache: dict = {}
    for r in res.data:
        if r.get("record_id"):
            rid_cache[r["record_id"].strip().lower()] = r["record_id"]
        if r.get("file_no"):
            fn_cache[r["file_no"].strip().lower()] = r["record_id"]
    return rid_cache, fn_cache


def build_closure_dedup_cache(sb, workspace_id: str) -> tuple[dict, dict]:
    """
    Return existing closure records for dedup:
      by_record_id: {record_id_lower: db_row}
      by_file_no:   {file_no_lower: db_row}
    """
    res = (
        sb.table("dggi_closure_records")
        .select("id,record_id,file_no")
        .eq("workspace_id", workspace_id)
        .execute()
    )
    by_rid: dict = {}
    by_fn: dict = {}
    for r in res.data:
        if r.get("record_id"):
            by_rid[r["record_id"].strip().lower()] = r
        if r.get("file_no"):
            by_fn[r["file_no"].strip().lower()] = r
    return by_rid, by_fn


# ---------------------------------------------------------------------------
# Upsert
# ---------------------------------------------------------------------------

def upsert_row(
    sb,
    workspace_id: str,
    sr_no: str,
    payload: dict,
    by_rid: dict,
    by_fn: dict,
    skipped: list,
    log: list,
    dry_run: bool,
) -> str:
    record_id = payload["record_id"]
    file_no = payload.get("file_no")

    try:
        existing = None
        match_by = None

        # 1. Exact record_id match
        if record_id and record_id.lower() in by_rid:
            existing = by_rid[record_id.lower()]
            match_by = "record_id"

        # 2. file_no fallback
        elif file_no and file_no.lower() in by_fn:
            existing = by_fn[file_no.lower()]
            match_by = "file_no"

        if existing:
            db_id = existing["id"]
            old_rid = existing["record_id"]
            if dry_run:
                print(f"    → UPDATE ({match_by})  existing={old_rid!r} → {record_id!r}  db_id={db_id}")
            else:
                sb.table("dggi_closure_records").update(payload).eq("id", db_id).execute()
            log.append({
                "action": "update",
                "match_by": match_by,
                "sr_no": sr_no,
                "old_record_id": old_rid,
                "new_record_id": record_id,
                "db_id": db_id,
            })
            return "updated"

        # 3. Insert
        insert_payload = {**payload, "workspace_id": workspace_id}
        if dry_run:
            print(f"    → INSERT  record_id={record_id!r}")
            inserted_id = None
        else:
            res = sb.table("dggi_closure_records").insert(insert_payload).execute()
            inserted_id = res.data[0]["id"] if res.data else None
        log.append({
            "action": "insert",
            "sr_no": sr_no,
            "new_record_id": record_id,
            "db_id": inserted_id,
        })
        return "inserted"

    except Exception as e:
        skipped.append({"sr_no": sr_no, "record_id": record_id, "reason": str(e)})
        print(f"    ERROR sr_no={sr_no}: {e}")
        return "skipped"


# ---------------------------------------------------------------------------
# Sheet processors
# ---------------------------------------------------------------------------

def process_fp_sheet(ws, sb, workspace_id, user_cache, rid_cache, fn_cache,
                     by_rid, by_fn, skipped, log, dry_run):
    """FP sheet: Full Payment closures."""
    print("\n--- Processing sheet: FP (Full Payment) ---")
    rows = list(ws.iter_rows(min_row=2, values_only=True))  # skip header row 1

    inserted = updated = skipped_count = 0
    unmatched_officers: set = set()
    unmatched_ir_nos: set = set()

    for row in rows:
        if all(v is None for v in row):
            continue
        sr_no = clean(row[0]) or "?"
        file_no = clean(row[1]) if len(row) > 1 else None
        taxpayer = clean(row[2]) if len(row) > 2 else None
        record_id = clean(row[3]) if len(row) > 3 else None
        closure_date = parse_date(row[4]) if len(row) > 4 else None
        section = clean(row[5]) if len(row) > 5 else None
        ir_no_raw = clean(row[6]) if len(row) > 6 else None
        payment_raw = clean(row[7]) if len(row) > 7 else None
        group_raw = clean(row[8]) if len(row) > 8 else None
        sio_raw = clean(row[9]) if len(row) > 9 else None

        if not record_id and not taxpayer:
            continue

        group = normalize_group(group_raw)

        # Resolve SIO → UUID
        handling_io_sio_id = None
        if sio_raw:
            handling_io_sio_id = user_cache.get(sio_raw.strip().lower())
            if not handling_io_sio_id:
                unmatched_officers.add(sio_raw)

        # Resolve source_record_id: IR No. first, file_no fallback
        source_record_id = None
        if ir_no_raw:
            source_record_id = rid_cache.get(ir_no_raw.strip().lower())
        if not source_record_id and file_no:
            source_record_id = fn_cache.get(file_no.strip().lower())
        if ir_no_raw and not source_record_id:
            unmatched_ir_nos.add(ir_no_raw)

        total_recovery = parse_amount_to_int(payment_raw)

        ALWAYS = {"record_id", "is_ir", "closure_by"}
        payload = {k: v for k, v in {
            "record_id": record_id,
            "source_record_id": source_record_id,
            "is_ir": True,
            "file_no": file_no,
            "taxpayer_name": taxpayer,
            "due_date": closure_date,
            "issue_involved": section,
            "total_recovery": total_recovery,
            "group": group,
            "handling_io_sio": handling_io_sio_id,
            "closure_by": "Closed After Payment of Tax",
        }.items() if v is not None or k in ALWAYS}

        if dry_run:
            print(
                f"  [#{sr_no}] record_id={record_id!r}\n"
                f"          file_no={file_no!r} | taxpayer={taxpayer!r}\n"
                f"          date={closure_date} | group={group!r}\n"
                f"          sio={sio_raw!r} → {handling_io_sio_id or 'NO MATCH'}\n"
                f"          ir_no={ir_no_raw!r} → src={source_record_id or 'NO MATCH'}\n"
                f"          payment={payment_raw!r} → {total_recovery}"
            )

        result = upsert_row(sb, workspace_id, sr_no, payload, by_rid, by_fn, skipped, log, dry_run)
        if result == "inserted":
            inserted += 1
        elif result == "updated":
            updated += 1
        else:
            skipped_count += 1

    print(
        f"\n[FP]  {'Would insert' if dry_run else 'Inserted'}: {inserted}"
        f" | {'Would update' if dry_run else 'Updated'}: {updated}"
        f" | Skipped: {skipped_count}"
    )
    if unmatched_officers:
        print(f"\nWarning: {len(unmatched_officers)} SIO name(s) not matched (handling_io_sio left null):")
        for n in sorted(unmatched_officers):
            print(f"  {n!r}")
    if unmatched_ir_nos:
        print(f"\nNote: {len(unmatched_ir_nos)} IR No(s) not found in dggi_records:")
        for ir in sorted(unmatched_ir_nos):
            print(f"  {ir!r}")


def process_nsp_sheet(ws, sb, workspace_id, user_cache, rid_cache, fn_cache,
                      by_rid, by_fn, skipped, log, dry_run):
    """NSP sheet: Non-Standard-Payment closures."""
    print("\n--- Processing sheet: NSP (Non-Standard Payment) ---")
    rows = list(ws.iter_rows(min_row=2, values_only=True))

    inserted = updated = skipped_count = 0
    unmatched_officers: set = set()

    for row in rows:
        if all(v is None for v in row):
            continue
        sr_no = clean(row[0]) or "?"
        file_no = clean(row[1]) if len(row) > 1 else None
        taxpayer = clean(row[2]) if len(row) > 2 else None
        gstins = clean(row[3]) if len(row) > 3 else None
        record_id = clean(row[4]) if len(row) > 4 else None
        closure_date = parse_date(row[5]) if len(row) > 5 else None
        section = clean(row[6]) if len(row) > 6 else None
        remark = clean(row[7]) if len(row) > 7 else None
        group_raw = clean(row[8]) if len(row) > 8 else None
        sio_raw = clean(row[9]) if len(row) > 9 else None

        if not record_id and not taxpayer:
            continue

        group = normalize_group(group_raw)

        # Resolve SIO → UUID
        handling_io_sio_id = None
        if sio_raw:
            handling_io_sio_id = user_cache.get(sio_raw.strip().lower())
            if not handling_io_sio_id:
                unmatched_officers.add(sio_raw)

        # Resolve source_record_id via file_no
        source_record_id = None
        if file_no:
            source_record_id = fn_cache.get(file_no.strip().lower())

        ALWAYS = {"record_id", "is_ir", "closure_by"}
        payload = {k: v for k, v in {
            "record_id": record_id,
            "source_record_id": source_record_id,
            "is_ir": True,
            "file_no": file_no,
            "taxpayer_name": taxpayer,
            "gstins": gstins,
            "due_date": closure_date,
            "issue_involved": section,
            "closure_reason": remark,
            "group": group,
            "handling_io_sio": handling_io_sio_id,
            "closure_by": "On Merit",
        }.items() if v is not None or k in ALWAYS}

        if dry_run:
            print(
                f"  [#{sr_no}] record_id={record_id!r}\n"
                f"          file_no={file_no!r} | taxpayer={taxpayer!r}\n"
                f"          gstin={gstins!r} | date={closure_date} | group={group!r}\n"
                f"          sio={sio_raw!r} → {handling_io_sio_id or 'NO MATCH'}\n"
                f"          remark={remark!r}"
            )

        result = upsert_row(sb, workspace_id, sr_no, payload, by_rid, by_fn, skipped, log, dry_run)
        if result == "inserted":
            inserted += 1
        elif result == "updated":
            updated += 1
        else:
            skipped_count += 1

    print(
        f"\n[NSP]  {'Would insert' if dry_run else 'Inserted'}: {inserted}"
        f" | {'Would update' if dry_run else 'Updated'}: {updated}"
        f" | Skipped: {skipped_count}"
    )
    if unmatched_officers:
        print(f"\nWarning: {len(unmatched_officers)} SIO name(s) not matched (handling_io_sio left null):")
        for n in sorted(unmatched_officers):
            print(f"  {n!r}")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    args = sys.argv[1:]
    dry_run = "--dry-run" in args
    positional = [a for a in args if not a.startswith("--")]
    excel_path = os.path.abspath(positional[0] if positional else DEFAULT_EXCEL_PATH)

    if not os.path.exists(excel_path):
        raise SystemExit(f"Excel file not found: {excel_path}")

    if dry_run:
        print("*** DRY RUN — no changes will be written to the database ***\n")

    print(f"Loading: {excel_path}")
    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
    print(f"Sheets found: {wb.sheetnames}")

    sb = create_client(SUPABASE_URL, SERVICE_ROLE_KEY)

    res = (
        sb.table("votum_users")
        .select("workspace_id")
        .eq("email", WORKSPACE_OWNER_EMAIL)
        .limit(1)
        .execute()
    )
    if not res.data:
        raise SystemExit(f"No user found for {WORKSPACE_OWNER_EMAIL!r}")
    workspace_id = res.data[0]["workspace_id"]
    print(f"Workspace: {workspace_id}")

    user_cache = build_user_cache(sb, workspace_id)
    print(f"Loaded {len(user_cache)} users")

    rid_cache, fn_cache = build_source_cache(sb, workspace_id)
    print(f"Loaded {len(rid_cache)} case records by record_id, {len(fn_cache)} by file_no")

    by_rid, by_fn = build_closure_dedup_cache(sb, workspace_id)
    print(f"Loaded {len(by_rid)} existing closure records by record_id, {len(by_fn)} by file_no")

    skipped: list = []
    log: list = []

    if "FP" in wb.sheetnames:
        process_fp_sheet(
            wb["FP"], sb, workspace_id,
            user_cache, rid_cache, fn_cache,
            by_rid, by_fn, skipped, log, dry_run,
        )
    else:
        print("Warning: Sheet 'FP' not found, skipping.")

    if "NSP" in wb.sheetnames:
        process_nsp_sheet(
            wb["NSP"], sb, workspace_id,
            user_cache, rid_cache, fn_cache,
            by_rid, by_fn, skipped, log, dry_run,
        )
    else:
        print("Warning: Sheet 'NSP' not found, skipping.")

    if log and not dry_run:
        with open(LOG_JSON, "w") as f:
            json.dump(log, f, indent=2)
        inserts = sum(1 for e in log if e["action"] == "insert")
        updates = sum(1 for e in log if e["action"] == "update")
        print(f"\nLog written to {LOG_JSON}  ({inserts} inserts, {updates} updates)")

    if skipped and not dry_run:
        with open(SKIPPED_CSV, "w", newline="") as f:
            writer = csv.DictWriter(f, fieldnames=list(skipped[0].keys()))
            writer.writeheader()
            writer.writerows(skipped)
        print(f"Skipped {len(skipped)} rows — see {SKIPPED_CSV}")
    elif skipped:
        print(f"\nWould skip {len(skipped)} rows (errors during DB lookup).")
    else:
        print("\nNo rows skipped.")

    print("\nDry run complete — no data written." if dry_run else "\nDone.")


if __name__ == "__main__":
    main()
