"""
Ingest IR pending cases data into Supabase dggi_records.

Source: 'MZU_ Pending IR cases.xlsx'  (sheet '30062026', ~92 rows)

Column mapping:
  Col 0  Sr. No.                            → upsert key (fallback record_id = "DIGIT-{sr_no:03d}")
  Col 1  Name of the Taxpayer               → taxpayer_name
  Col 2  GSTIN/PAN                          → gstins
  Col 3  IR No./335-J No.                   → upsert key 1 (exact match vs record_id)
  Col 4  Date of Detection/IR Date          → date_of_ir + date_of_initiation
  Col 5  (second date, skip)
  Col 6  Pending since (days)               → (skipped)
  Col 7  Pendency Year Wise                 → (skipped)
  Col 8  Detection (in Lakhs)               → detection_amount (× 1,00,000)
  Col 9  Additional Detection (In Lakhs)    → (skipped)
  Col 10 Recovery (In Lakh)                 → recovery_itc (× 1,00,000)
  Col 11 Additional Recovery (In Lakhs)     → (skipped)
  Col 12 Type of Case                       → (skipped)
  Col 13 Brief Facts of the Case            → issue_involved
  Col 14 Present Status                     → latest_status
  Col 15 Expected Date of Closure/SCN       → (skipped)
  Col 16 Name of SIO                        → sio_name
  Col 17 Group                              → group (prefixed "Group " + letter)
  Col 18 Section (73/74/76)                 → (skipped)
  Col 19 Whether in DIGIT (DIGIT No.)       → digit_id (upsert key 2; only real DIGIT IDs kept)

Dedup strategy (checked in order):
  1. IR No. match    — exact match of col 3 vs record_id in DB  (primary)
  2. digit_id match  — col 19 numeric-format IDs matched vs DB digit_id  (fallback)
  3. No match        → insert; record_id = IR No. (fallback "DIGIT-{sr_no:03d}")

Usage:
    python3 scripts/ingest_ir_digit_data.py [/path/to/file.xlsx] [--dry-run]
"""

import csv
import json
import os
import re
import sys
from datetime import date, datetime

import openpyxl
from dotenv import load_dotenv

from supabase import create_client

load_dotenv(os.path.join(os.path.dirname(__file__), ".env"))
SUPABASE_URL = os.environ["SUPABASE_URL"]
SERVICE_ROLE_KEY = os.environ["SERVICE_ROLE_KEY"]

WORKSPACE_OWNER_EMAIL = "ajinkya.k1@gov.in"

DEFAULT_EXCEL_PATH = os.path.join(
    os.path.dirname(__file__),
    "..",
    "data",
    "MZU_ Pending IR cases.xlsx",
)

SHEET_NAME = "04082026"

SKIPPED_CSV = os.path.join(os.path.dirname(__file__), "ingest_ir_digit_skipped.csv")
LOG_JSON = os.path.join(os.path.dirname(__file__), "ingest_ir_digit_log.json")


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def parse_date(val) -> str | None:
    if val is None:
        return None
    if isinstance(val, (datetime, date)):
        d = val.date() if isinstance(val, datetime) else val
        # Excel (US locale) sometimes stores DD/MM-entered dates with day and month swapped.
        # If the result is a future date and the values are swappable, correct it.
        if d > date.today() and d.day <= 12:
            try:
                d = d.replace(month=d.day, day=d.month)
            except ValueError:
                pass
        if d > date.today():
            return None
        return d.isoformat()
    s = str(val).strip()
    if not s:
        return None
    for fmt in ("%d-%m-%Y", "%d/%m/%Y", "%d.%m.%Y"):
        try:
            return datetime.strptime(s, fmt).date().isoformat()
        except ValueError:
            pass
    return None


def clean(val) -> str | None:
    if val is None:
        return None
    s = str(val).strip()
    return s if s else None


def str_amount(val) -> str | None:
    """Convert a value stored in lakhs to absolute rupees (× 1,00,000)."""
    if val is None:
        return None
    try:
        f = float(val)
        return str(f * 100_000) if f != 0 else None
    except (ValueError, TypeError):
        s = str(val).strip()
        return s if s else None


def strip_digit_prefix(digit_id: str | None) -> str | None:
    """'DIGIT-20260401132436-258' → '20260401132436-258' (how DB stores it)"""
    if not digit_id:
        return None
    if digit_id.upper().startswith("DIGIT-"):
        return digit_id[6:]
    return digit_id


def _is_real_digit_id(val: str | None) -> bool:
    """Return True only for numeric-format DIGIT IDs like '20251103132254-640'."""
    if not val:
        return False
    s = val.strip()
    if s.lower() in ("ok", "not in digit", ""):
        return False
    return bool(re.match(r"^\d{10,}-\d+$", s)) or bool(re.match(r"^DIGIT-\d", s, re.IGNORECASE))


# ---------------------------------------------------------------------------
# Upsert: IR No (record_id) → digit_id → insert
# ---------------------------------------------------------------------------

def upsert_ir_record(
    sb,
    workspace_id: str,
    sr_no: int,
    digit_id_raw: str | None,
    ir_no: str | None,
    payload: dict,
    skipped: list,
    log: list,
    dry_run: bool = False,
) -> str:
    try:
        # 1. Match on IR No. vs record_id (primary key)
        if ir_no:
            res = (
                sb.table("dggi_records")
                .select("id,record_id")
                .eq("workspace_id", workspace_id)
                .eq("record_id", ir_no)
                .execute()
            )
            if res.data:
                row = res.data[0]
                if dry_run:
                    print(f"    → UPDATE (ir_no)  existing record_id={row['record_id']!r}  db_id={row['id']}")
                else:
                    sb.table("dggi_records").update(payload).eq("id", row["id"]).execute()
                log.append({"action": "update", "match_by": "ir_no", "sr_no": sr_no, "excel_ir_no": ir_no, "existing_record_id": row["record_id"], "db_id": row["id"], "digit_id": digit_id_raw, "taxpayer_name": payload.get("taxpayer_name")})
                return "updated"

        # 2. Match on digit_id
        digit_id_db = strip_digit_prefix(digit_id_raw)
        if digit_id_db:
            res = (
                sb.table("dggi_records")
                .select("id,record_id")
                .eq("workspace_id", workspace_id)
                .eq("digit_id", digit_id_db)
                .execute()
            )
            if res.data:
                row = res.data[0]
                if dry_run:
                    print(f"    → UPDATE (digit_id)  existing record_id={row['record_id']!r}  db_id={row['id']}")
                else:
                    sb.table("dggi_records").update(payload).eq("id", row["id"]).execute()
                log.append({"action": "update", "match_by": "digit_id", "sr_no": sr_no, "excel_ir_no": ir_no, "existing_record_id": row["record_id"], "db_id": row["id"], "digit_id": digit_id_raw, "taxpayer_name": payload.get("taxpayer_name")})
                return "updated"

        # 3. Insert — use ir_no as record_id (fallback to DIGIT-{sr_no:03d} if missing)
        new_record_id = ir_no or f"DIGIT-{sr_no:03d}"
        insert_payload = {**payload, "workspace_id": workspace_id, "record_id": new_record_id}
        if dry_run:
            print(f"    → INSERT  new record_id={new_record_id!r}")
            inserted_id = None
        else:
            insert_res = sb.table("dggi_records").insert(insert_payload).execute()
            inserted_id = insert_res.data[0]["id"] if insert_res.data else None
        log.append({"action": "insert", "sr_no": sr_no, "new_record_id": new_record_id, "db_id": inserted_id, "excel_ir_no": ir_no, "digit_id": digit_id_raw, "taxpayer_name": payload.get("taxpayer_name")})
        return "inserted"

    except Exception as e:
        skipped.append({"sr_no": sr_no, "digit_id": digit_id_raw or "", "ir_no": ir_no or "", "reason": str(e)})
        return "skipped"


# ---------------------------------------------------------------------------
# Sheet processor
# ---------------------------------------------------------------------------

def process_sheet(ws, sb, workspace_id: str, skipped: list, log: list, dry_run: bool = False):
    rows = list(ws.iter_rows(values_only=True))
    inserted = updated = skipped_count = 0

    # Skip row 0 (title) and row 1 (header); data starts at row 2 (index 2)
    for row in rows[2:]:
        if row[0] is None:
            continue
        try:
            sr_no = int(row[0])
        except (ValueError, TypeError):
            continue

        taxpayer_name = clean(row[1])
        if not taxpayer_name:
            continue

        ir_date = parse_date(row[4])
        ir_no = clean(row[3])

        digit_id_raw_col19 = clean(row[19])
        digit_id_raw = digit_id_raw_col19 if _is_real_digit_id(digit_id_raw_col19) else None

        payload = {
            "taxpayer_name": taxpayer_name,
            "digit_id": strip_digit_prefix(digit_id_raw),
            "gstins": clean(row[2]),
            "date_of_ir": ir_date,
            "date_of_initiation": ir_date,
            "detection_amount": str_amount(row[8]),
            "recovery_itc": str_amount(row[10]),
            "latest_status": clean(row[14]),
            "issue_involved": clean(row[13]),
            "sio_name": clean(row[16]),
            "group": f"Group {clean(row[17])}" if clean(row[17]) else None,
            "is_ir": True,
        }
        payload = {k: v for k, v in payload.items() if v is not None}

        if dry_run:
            print(
                f"  [#{sr_no:02d}] {taxpayer_name[:45]!r}"
                f" | digit_id={digit_id_raw} | ir_no={ir_no}"
            )

        result = upsert_ir_record(
            sb, workspace_id, sr_no, digit_id_raw, ir_no, payload, skipped, log, dry_run
        )
        if result == "inserted":
            inserted += 1
        elif result == "updated":
            updated += 1
        else:
            skipped_count += 1

    print(
        f"\n[IR DIGIT]  {'Would insert' if dry_run else 'Inserted'}: {inserted}"
        f" | {'Would update' if dry_run else 'Updated'}: {updated}"
        f" | Skipped: {skipped_count}"
    )


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    args = sys.argv[1:]
    dry_run = "--dry-run" in args
    positional = [a for a in args if not a.startswith("--")]
    excel_path = os.path.abspath(positional[0] if positional else DEFAULT_EXCEL_PATH)

    if not os.path.exists(excel_path):
        raise SystemExit(f"Excel file not found: {excel_path}")

    if dry_run:
        print("*** DRY RUN — no changes will be written to the database ***\n")

    print(f"Loading: {excel_path}")
    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)

    if SHEET_NAME not in wb.sheetnames:
        raise SystemExit(f"Sheet {SHEET_NAME!r} not found. Available: {wb.sheetnames}")

    sb = create_client(SUPABASE_URL, SERVICE_ROLE_KEY)

    res = (
        sb.table("votum_users")
        .select("workspace_id")
        .eq("email", WORKSPACE_OWNER_EMAIL)
        .limit(1)
        .execute()
    )
    if not res.data:
        raise SystemExit(f"No user found for {WORKSPACE_OWNER_EMAIL}")
    workspace_id = res.data[0]["workspace_id"]
    print(f"Workspace: {workspace_id}\n")

    skipped = []
    log = []
    process_sheet(wb[SHEET_NAME], sb, workspace_id, skipped, log, dry_run)

    if log:
        with open(LOG_JSON, "w") as f:
            json.dump(log, f, indent=2)
        print(f"\nLog written to {LOG_JSON}  ({len([e for e in log if e['action']=='insert'])} inserts, {len([e for e in log if e['action']=='update'])} updates)")

    if skipped:
        if not dry_run:
            with open(SKIPPED_CSV, "w", newline="") as f:
                writer = csv.DictWriter(f, fieldnames=list(skipped[0].keys()))
                writer.writeheader()
                writer.writerows(skipped)
            print(f"\nSkipped {len(skipped)} rows — see {SKIPPED_CSV}")
        else:
            print(f"\nWould skip {len(skipped)} rows (errors during DB lookup).")
    else:
        print("\nNo rows skipped.")

    print("\nDry run complete — no data written." if dry_run else "\nDone.")


if __name__ == "__main__":
    main()
