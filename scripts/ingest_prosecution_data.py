"""
Ingest Prosecution List data into Supabase.

Source: 'Prosecution List DGGI MZU_Updated.xlsx'
  - Sheet 'Arrest-Prosecution'    → dggi_prosecution_arrest_records  (53 rows)
  - Sheet 'Non-Arrest Prosecution' → dggi_prosecution_non_arrest_records (25 rows)

Both sheets share the same column layout (0-based):
  Col 0  SR.NO
  Col 1  Group                              → group
  Col 2  SIO                                → sio_name
  Col 3  ACT                                → (informational, skipped)
  Col 4  F.Y.                               → (informational, skipped)
  Col 5  F. NO.                             → linked_case_id (lookup via dggi_records.file_no)
  Col 6  NAME OF THE TAXPAYER               → entity_name
  Col 7  ADDRESS                            → (skipped)
  Col 8  GSTIN                              → gstin
  Col 9  AMOUNT OF TAX EVADED (In Lakhs)    → amount_evaded_crore (÷100)
  Col 10 NAME OF THE PERSON ARRESTED/PROSECUTED → arrested_person_name / person_name
  Col 11 DESIGNATION                        → (skipped)
  Col 12 AGE                                → age
  Col 13 DATE OF ARREST (col 13 in Arrest, col 13 in Non-Arrest; both labelled same)
         → date_of_arrest
  Col 14 ROLE PLAYED                        → brief_modus_operandi
  Col 15 PROSECUTION SANCTION ORDER NO.     → (informational)
  Col 16 PROSECUTION SANCTION ORDER DATE    → (informational)
  Col 17 NUMBER OF PERSONS PROSECUTED       → (informational)
  Col 18 DATE OF FILING OF COMPLAINT        → date_of_filing
  Col 19 CRIMINAL COMPLAINT NUMBER          → prosecution_complaint_status prefix
  Col 20 ORDER NO.  (merged header row 2)   → (informational)
  Col 21 ORDER DATE (merged header row 2)   → (informational)
  Col 22 ORDER DETAILS (merged header row 2)→ (informational)
  Col 23 CURRENT STATUS / Next hearing      → prosecution_complaint_status
  Col 24 NAME OF THE COUNSEL                → (skipped)

Insert strategy: ALWAYS INSERT (no upsert). Re-running creates duplicates.
Record IDs: PRA/###/YY-YY for arrest prosecutions, PRN/###/YY-YY for non-arrest.

Usage:
    python3 scripts/ingest_prosecution_data.py [/path/to/file.xlsx] [--dry-run]
"""

import csv
import json
import os
import sys
from datetime import date, datetime

import openpyxl
from dotenv import load_dotenv
from supabase import create_client

load_dotenv(os.path.join(os.path.dirname(__file__), ".env"))
SUPABASE_URL = os.environ["SUPABASE_URL"]
SERVICE_ROLE_KEY = os.environ["SERVICE_ROLE_KEY"]

WORKSPACE_OWNER_EMAIL = "ajinkya.k1@gov.in"

DEFAULT_EXCEL_PATH = os.path.join(
    os.path.dirname(__file__),
    "..",
    "data",
    "Prosecution List DGGI MZU_Updated.xlsx",
)

SKIPPED_CSV = os.path.join(os.path.dirname(__file__), "ingest_prosecution_skipped.csv")
LOG_JSON = os.path.join(os.path.dirname(__file__), "ingest_prosecution_log.json")


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def parse_date(val) -> str | None:
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.date().isoformat()
    if isinstance(val, date):
        return val.isoformat()
    s = str(val).strip()
    if not s or s.upper() == "NA":
        return None
    for fmt in ("%d.%m.%Y", "%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(s, fmt).date().isoformat()
        except ValueError:
            pass
    return None


def clean(val) -> str | None:
    if val is None:
        return None
    s = str(val).strip()
    if s == "-" or s.upper() == "NA" or s == "":
        return None
    return s


def lakhs_to_crores(val) -> str | None:
    """Convert amount in lakhs to crores (÷100)."""
    if val is None:
        return None
    try:
        lakhs = float(val)
        if lakhs == 0:
            return None
        return str(round(lakhs / 100, 2))
    except (ValueError, TypeError):
        return None


def normalize_group(val) -> str | None:
    if val is None:
        return None
    s = str(val).strip()
    if not s:
        return None
    if s.startswith("Group "):
        return s
    if len(s) == 1 and s.upper() in "ABCDEF":
        return f"Group {s.upper()}"
    return s


def current_fy() -> str:
    """Returns FY in YY-YY format, e.g. '26-27'."""
    now = datetime.now()
    yr = now.year
    start = yr if now.month >= 4 else yr - 1
    return f"{str(start)[2:]}-{str(start + 1)[2:]}"


def lookup_linked_case_id(sb, workspace_id: str, file_no: str | None) -> str | None:
    if not file_no:
        return None
    res = (
        sb.table("dggi_records")
        .select("record_id")
        .eq("workspace_id", workspace_id)
        .eq("file_no", file_no)
        .limit(1)
        .execute()
    )
    return res.data[0]["record_id"] if res.data else None


def next_seq(sb, workspace_id: str, table: str, prefix: str) -> int:
    import re
    res = (
        sb.table(table)
        .select("record_id")
        .eq("workspace_id", workspace_id)
        .like("record_id", f"{prefix}/%")
        .execute()
    )
    max_seq = 0
    for r in res.data:
        m = re.match(rf"{re.escape(prefix)}/(\d+)/", r["record_id"] or "")
        if m:
            max_seq = max(max_seq, int(m.group(1)))
    return max_seq + 1


# ---------------------------------------------------------------------------
# Insert functions
# ---------------------------------------------------------------------------

def insert_arrest_record(
    sb,
    workspace_id: str,
    arr_seq: int,
    row: tuple,
    linked_case: str | None,
    skipped: list,
    log: list,
    dry_run: bool,
) -> tuple[str, str | None]:
    """Insert one row into dggi_arrest_records. Returns (status, arrest_uuid)."""
    sr_no = row[0]
    fy = current_fy()
    record_id = f"ARR/{arr_seq:03d}/{fy}"

    party_name = clean(row[6]) or ""
    payload = {
        "workspace_id": workspace_id,
        "record_id": record_id,
        "linked_case_id": linked_case,
        "arrest_batch_id": clean(row[5]),  # F. NO. groups all persons from the same case
        "group": normalize_group(row[1]),
        "sio_name": clean(row[2]),
        "financial_year": clean(row[4]),
        "party_name": party_name,
        "unit_gstin": clean(row[8]) or "",
        "amount_crore": lakhs_to_crores(row[9]),
        "arrested_name": clean(row[10]),
        "arrested_designation": clean(row[11]),
        "arrested_age": clean(row[12]),
        "date_of_arrest": parse_date(row[13]),
        "role_evidence": clean(row[14]),
        "prosecution_filed": "Yes" if (clean(row[15]) or parse_date(row[18])) else "",
    }
    payload = {k: v for k, v in payload.items() if v is not None}

    if dry_run:
        print(f"    → INSERT ARR  record_id={record_id!r}  name={payload.get('arrested_name')!r}")
        return "inserted", "dry-run-uuid"

    try:
        res = sb.table("dggi_arrest_records").insert(payload).execute()
        arrest_uuid = res.data[0]["id"] if res.data else None
    except Exception as e:
        skipped.append({"sheet": "Arrest-Prosecution", "sr_no": sr_no, "record_id": record_id, "reason": f"arrest insert: {e}"})
        return "skipped", None

    log.append({"action": "insert_arrest", "sheet": "Arrest-Prosecution", "sr_no": sr_no, "record_id": record_id, "arrested_name": payload.get("arrested_name"), "db_id": arrest_uuid})
    return "inserted", arrest_uuid


def insert_arrest_prosecution(
    sb,
    workspace_id: str,
    pra_seq: int,
    row: tuple,
    linked_case: str | None,
    arrest_uuid: str | None,
    skipped: list,
    log: list,
    dry_run: bool,
) -> str:
    """Insert one row into dggi_prosecution_arrest_records, linked to its arrest record."""
    sr_no = row[0]
    person_name = clean(row[10])
    if not person_name:
        skipped.append({"sheet": "Arrest-Prosecution", "sr_no": sr_no, "reason": "missing person name"})
        return "skipped"

    fy = current_fy()
    record_id = f"PRA/{pra_seq:03d}/{fy}"

    payload = {
        "workspace_id": workspace_id,
        "record_id": record_id,
        "linked_arrest_id": arrest_uuid,
        "linked_case_id": linked_case,
        "group": normalize_group(row[1]),
        "sio_name": clean(row[2]),
        "entity_name": clean(row[6]),
        "gstin": clean(row[8]),
        "amount_evaded_crore": lakhs_to_crores(row[9]),
        "arrested_person_name": person_name,
        "age": clean(row[12]),
        "date_of_arrest": parse_date(row[13]),
        "brief_modus_operandi": clean(row[14]),
        "date_of_filing": parse_date(row[18]),
        "prosecution_complaint_status": clean(row[23]),
    }
    payload = {k: v for k, v in payload.items() if v is not None}

    if dry_run:
        print(f"    → INSERT PRA  record_id={record_id!r}  linked_arrest={arrest_uuid!r}")
    else:
        try:
            sb.table("dggi_prosecution_arrest_records").insert(payload).execute()
        except Exception as e:
            skipped.append({"sheet": "Arrest-Prosecution", "sr_no": sr_no, "record_id": record_id, "reason": f"prosecution insert: {e}"})
            return "skipped"

    log.append({"action": "insert_prosecution", "sheet": "Arrest-Prosecution", "sr_no": sr_no, "record_id": record_id, "person_name": person_name, "linked_arrest_id": arrest_uuid})
    return "inserted"


def insert_non_arrest_prosecution(
    sb,
    workspace_id: str,
    seq: int,
    sheet_name: str,
    row: tuple,
    skipped: list,
    log: list,
    dry_run: bool,
) -> str:
    """Insert one row into dggi_prosecution_non_arrest_records."""
    sr_no = row[0]
    person_name = clean(row[10])
    if not person_name:
        skipped.append({"sheet": sheet_name, "sr_no": sr_no, "reason": "missing person name"})
        return "skipped"

    fy = current_fy()
    record_id = f"PRN/{seq:03d}/{fy}"
    file_no = clean(row[5])

    linked_case = None if dry_run else lookup_linked_case_id(sb, workspace_id, file_no)

    payload = {
        "workspace_id": workspace_id,
        "record_id": record_id,
        "linked_case_id": linked_case,
        "group": normalize_group(row[1]),
        "sio_name": clean(row[2]),
        "entity_name": clean(row[6]),
        "gstin": clean(row[8]),
        "amount_evaded_crore": lakhs_to_crores(row[9]),
        "person_name": person_name,
        "age": clean(row[12]),
        "date_of_arrest": parse_date(row[13]),
        "brief_modus_operandi": clean(row[14]),
        "date_of_filing": parse_date(row[18]),
        "prosecution_complaint_status": clean(row[23]),
    }
    payload = {k: v for k, v in payload.items() if v is not None}

    if dry_run:
        print(f"    → INSERT PRN  record_id={record_id!r}  person={person_name!r}")
    else:
        try:
            sb.table("dggi_prosecution_non_arrest_records").insert(payload).execute()
        except Exception as e:
            skipped.append({"sheet": sheet_name, "sr_no": sr_no, "record_id": record_id, "reason": str(e)})
            return "skipped"

    log.append({"action": "insert", "sheet": sheet_name, "sr_no": sr_no, "record_id": record_id, "person_name": person_name})
    return "inserted"


# ---------------------------------------------------------------------------
# Sheet processors
# ---------------------------------------------------------------------------

def process_arrest_sheet(ws, sb, workspace_id: str, skipped: list, log: list, dry_run: bool) -> tuple[int, int]:
    rows = list(ws.iter_rows(values_only=True))
    inserted = skipped_count = 0
    arr_seq = 1 if dry_run else next_seq(sb, workspace_id, "dggi_arrest_records", "ARR")
    pra_seq = 1 if dry_run else next_seq(sb, workspace_id, "dggi_prosecution_arrest_records", "PRA")

    for row in rows[2:]:  # rows 0-1 are header
        if row[0] is None:
            continue
        try:
            int(row[0])
        except (ValueError, TypeError):
            continue

        if dry_run:
            print(f"  [Arrest #{row[0]}] {str(row[10] or '')[:45]!r}")

        file_no = clean(row[5])
        linked_case = None if dry_run else lookup_linked_case_id(sb, workspace_id, file_no)

        # 1. Insert arrest record
        arr_status, arrest_uuid = insert_arrest_record(
            sb, workspace_id, arr_seq, row, linked_case, skipped, log, dry_run
        )
        if arr_status == "inserted":
            arr_seq += 1
        else:
            skipped_count += 1
            continue  # skip prosecution if arrest failed

        # 2. Insert linked prosecution record
        pra_status = insert_arrest_prosecution(
            sb, workspace_id, pra_seq, row, linked_case, arrest_uuid, skipped, log, dry_run
        )
        if pra_status == "inserted":
            inserted += 1
            pra_seq += 1
        else:
            skipped_count += 1

    print(
        f"\n[Arrest-Prosecution]  {'Would insert' if dry_run else 'Inserted'}: {inserted} prosecution records"
        f" | Skipped: {skipped_count}"
    )
    return inserted, skipped_count


def process_non_arrest_sheet(ws, sheet_name: str, sb, workspace_id: str, skipped: list, log: list, dry_run: bool, seq_start: int) -> tuple[int, int]:
    rows = list(ws.iter_rows(values_only=True))
    inserted = skipped_count = 0
    seq = seq_start

    for row in rows[2:]:  # rows 0-1 are header
        if row[0] is None:
            continue
        try:
            int(row[0])
        except (ValueError, TypeError):
            continue

        if dry_run:
            print(f"  [{sheet_name} #{row[0]}] {str(row[10] or '')[:45]!r}")

        result = insert_non_arrest_prosecution(sb, workspace_id, seq, sheet_name, row, skipped, log, dry_run)
        if result == "inserted":
            inserted += 1
            seq += 1
        else:
            skipped_count += 1

    print(
        f"\n[{sheet_name}]  {'Would insert' if dry_run else 'Inserted'}: {inserted}"
        f" | Skipped: {skipped_count}"
    )
    return inserted, skipped_count


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    args = sys.argv[1:]
    dry_run = "--dry-run" in args
    positional = [a for a in args if not a.startswith("--")]
    excel_path = os.path.abspath(positional[0] if positional else DEFAULT_EXCEL_PATH)

    if not os.path.exists(excel_path):
        raise SystemExit(f"Excel file not found: {excel_path}")

    if dry_run:
        print("*** DRY RUN — no changes will be written to the database ***\n")

    print(f"Loading: {excel_path}")
    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)

    sb = create_client(SUPABASE_URL, SERVICE_ROLE_KEY)

    res = (
        sb.table("votum_users")
        .select("workspace_id")
        .eq("email", WORKSPACE_OWNER_EMAIL)
        .limit(1)
        .execute()
    )
    if not res.data:
        raise SystemExit(f"No user found for {WORKSPACE_OWNER_EMAIL}")
    workspace_id = res.data[0]["workspace_id"]
    print(f"Workspace: {workspace_id}\n")

    skipped: list = []
    log: list = []

    # --- Arrest-Prosecution sheet ---
    print("=== Sheet: Arrest-Prosecution ===")
    process_arrest_sheet(wb["Arrest-Prosecution"], sb, workspace_id, skipped, log, dry_run)

    # --- Non-Arrest Prosecution sheet ---
    prn_seq_start = 1 if dry_run else next_seq(sb, workspace_id, "dggi_prosecution_non_arrest_records", "PRN")

    print("\n=== Sheet: Non-Arrest Prosecution ===")
    process_non_arrest_sheet(
        wb["Non-Arrest Prosecution"], "Non-Arrest Prosecution",
        sb, workspace_id, skipped, log, dry_run, prn_seq_start
    )

    if log:
        with open(LOG_JSON, "w") as f:
            json.dump(log, f, indent=2)
        n_ins = len([e for e in log if e["action"] == "insert"])
        print(f"\nLog written to {LOG_JSON}  ({n_ins} inserts)")

    if skipped:
        if not dry_run:
            with open(SKIPPED_CSV, "w", newline="") as f:
                writer = csv.DictWriter(f, fieldnames=list(skipped[0].keys()))
                writer.writeheader()
                writer.writerows(skipped)
            print(f"\nSkipped {len(skipped)} rows — see {SKIPPED_CSV}")
        else:
            print(f"\nWould skip {len(skipped)} rows (errors during DB lookup).")
    else:
        print("\nNo rows skipped.")

    print("\nDry run complete — no data written." if dry_run else "\nDone.")


if __name__ == "__main__":
    main()
