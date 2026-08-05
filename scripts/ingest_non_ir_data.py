"""
Ingest NON-IR pending cases data into Supabase dggi_records.

Supports two Excel layouts (auto-detected from header row):

Layout A — original ('MZU_ Non-IR Pending.xlsx', sheet 'Pending Non-IR'):
  Col 0  Non-IR Register No. → legacy_non_ir_no
  Col 1  Name of TP          → taxpayer_name
  Col 2  GSTIN               → gstins
  Col 3  Initiation date     → date_of_non_ir (used for FY sequencing)
  Col 4  REIC/Co-lending     → (skipped)
  Col 5  SIO                 → sio_name
  Col 6  Group               → group (prefixed "Group " + letter)
  Col 7  Mode                → mode_of_initiation
  Col 8  Current Status      → latest_status
  Col 9  email SIO           → sio_email

Layout B — new ('MZU_ Non-IR Pending (70).xlsx', sheet 'Pending Non-IR (70)'):
  Col 0  Sr. No.             → (skipped)
  Col 1  Non-IR Register No. → legacy_non_ir_no
  Col 2  Name of TP          → taxpayer_name
  Col 3  GSTIN               → gstins
  Col 4  Initiation date     → date_of_non_ir (used for FY sequencing)
  Col 5  REIC/Co-lending     → (skipped)
  Col 6  SIO                 → sio_name
  Col 7  Group               → group (prefixed "Group " + letter)
  Col 8  Mode                → mode_of_initiation
  Col 9  Current Status      → latest_status
  (no email column)

record_id generation (FY-based, date-sequenced):
  - Extract short FY from date_of_non_ir (e.g. 2026-05-01 → "26-27")
  - Sort all records by date within each FY
  - Assign sequential numbers starting from 1: NIR-001-26-27, NIR-002-26-27, etc.
  - Sequence resets for each FY

Dedup strategy (checked in order):
  1. legacy_non_ir_no match — update existing row when col 1 is not null
  2. No match              → insert with date-sequenced record_id

Usage:
    python3 scripts/ingest_non_ir_data.py [/path/to/file.xlsx] [--dry-run]
"""

import csv
import json
import os
import sys
from collections import defaultdict
from datetime import date, datetime

import openpyxl
from dotenv import load_dotenv
from supabase import create_client

load_dotenv(os.path.join(os.path.dirname(__file__), ".env"))
SUPABASE_URL = os.environ["SUPABASE_URL"]
SERVICE_ROLE_KEY = os.environ["SERVICE_ROLE_KEY"]

WORKSPACE_OWNER_EMAIL = "ajinkya.k1@gov.in"

DEFAULT_EXCEL_PATH = os.path.join(
    os.path.dirname(__file__),
    "..",
    "data",
    "MZU_ Non-IR Pending (70).xlsx",
)

# Both sheet names accepted; the correct one is selected at runtime
SHEET_NAMES = ["Pending Non-IR (70)", "Pending Non-IR"]

SKIPPED_CSV = os.path.join(os.path.dirname(__file__), "ingest_non_ir_skipped.csv")
LOG_JSON = os.path.join(os.path.dirname(__file__), "ingest_non_ir_log.json")


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def parse_date(val) -> str | None:
    if val is None:
        return None
    if isinstance(val, (datetime, date)):
        d = val.date() if isinstance(val, datetime) else val
        if d > date.today() and d.day <= 12:
            try:
                d = d.replace(month=d.day, day=d.month)
            except ValueError:
                pass
        if d > date.today():
            return None
        return d.isoformat()
    s = str(val).strip()
    if not s:
        return None
    for fmt in ("%d.%m.%Y", "%d/%m/%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(s, fmt).date().isoformat()
        except ValueError:
            pass
    return None


def clean(val) -> str | None:
    if val is None:
        return None
    s = str(val).strip()
    return s if s else None


# ---------------------------------------------------------------------------
# FY helper
# ---------------------------------------------------------------------------

def fy_from_date(date_str: str | None) -> str:
    """'2026-05-01' → '26-27'; '2025-03-15' → '24-25'"""
    if not date_str:
        now = date.today()
        year = now.year if now.month >= 4 else now.year - 1
    else:
        year = int(date_str[:4])
        month = int(date_str[5:7])
        year = year if month >= 4 else year - 1
    return f"{year - 2000:02d}-{year + 1 - 2000:02d}"


# ---------------------------------------------------------------------------
# Batch email → votum_users.id resolver
# ---------------------------------------------------------------------------

def resolve_sio_emails(sb, workspace_id: str, emails: list[str]) -> dict[str, str]:
    """Return {email: user_id} for all emails found in votum_users."""
    unique = [e for e in set(emails) if e]
    if not unique:
        return {}
    res = (
        sb.table("votum_users")
        .select("id,email")
        .eq("workspace_id", workspace_id)
        .in_("email", unique)
        .execute()
    )
    return {row["email"]: row["id"] for row in (res.data or [])}


# ---------------------------------------------------------------------------
# Upsert: legacy_non_ir_no match → update; no match → insert
# ---------------------------------------------------------------------------

def upsert_non_ir_record(
    sb,
    workspace_id: str,
    sr_no: int,
    payload: dict,
    skipped: list,
    log: list,
    dry_run: bool = False,
) -> str:
    legacy_non_ir_no = payload.get("legacy_non_ir_no")
    record_id = payload["record_id"]
    try:
        # 1. Match on legacy_non_ir_no
        if legacy_non_ir_no:
            res = (
                sb.table("dggi_records")
                .select("id,record_id")
                .eq("workspace_id", workspace_id)
                .eq("legacy_non_ir_no", legacy_non_ir_no)
                .execute()
            )
            if res.data:
                row = res.data[0]
                if dry_run:
                    print(f"    → UPDATE (legacy_non_ir_no={legacy_non_ir_no!r})  existing={row['record_id']!r}  db_id={row['id']}")
                else:
                    sb.table("dggi_records").update(payload).eq("id", row["id"]).execute()
                log.append({"action": "update", "match_by": "legacy_non_ir_no", "sr_no": sr_no, "record_id": record_id, "db_id": row["id"]})
                return "updated"

        # 2. Insert with date-sequenced record_id
        insert_payload = {**payload, "workspace_id": workspace_id}
        if dry_run:
            print(f"    → INSERT  new record_id={record_id!r}")
            inserted_id = None
        else:
            insert_res = sb.table("dggi_records").insert(insert_payload).execute()
            inserted_id = insert_res.data[0]["id"] if insert_res.data else None
        log.append({"action": "insert", "sr_no": sr_no, "new_record_id": record_id, "db_id": inserted_id})
        return "inserted"

    except Exception as e:
        skipped.append({"sr_no": sr_no, "record_id": record_id, "reason": str(e)})
        return "skipped"


# ---------------------------------------------------------------------------
# Sheet processor — groups by FY, sorts by date, assigns record_ids
# ---------------------------------------------------------------------------

def _detect_col_offsets(header_row) -> tuple[int, bool]:
    """
    Return (offset, has_email) based on the header row.

    Layout B has 'Sr. No.' in col 0, shifting every column right by 1.
    Layout A starts directly with 'Non-IR Register No.' in col 0.
    """
    first = str(header_row[0] or "").lower()
    if "sr" in first or "no" in first.split():
        # Layout B: Sr. No. prepended
        return 1, False
    # Layout A: original layout, email in col 9
    return 0, True


def process_sheet(ws, sb, workspace_id: str, skipped: list, log: list, dry_run: bool = False):
    rows = list(ws.iter_rows(values_only=True))

    # Skip row 0 (title) and row 1 (header); data starts at index 2
    header_row = rows[1]
    offset, has_email = _detect_col_offsets(header_row)
    print(f"  Column layout: {'B (Sr.No. prefix, no email)' if offset else 'A (original, with email)'}")

    raw_records = []
    for idx, row in enumerate(rows[2:], start=1):
        # Skip blank rows: both the non-ir-no cell and the name cell are None
        if row[offset] is None and row[offset + 1] is None:
            continue

        nir_date = parse_date(row[offset + 3]) or date.today().isoformat()
        group_raw = clean(row[offset + 6])
        officer_name = clean(row[offset + 5])

        raw_records.append({
            "sr_no": idx,
            "legacy_non_ir_no": clean(row[offset]),
            "date": nir_date,
            "taxpayer_name": clean(row[offset + 1]) or officer_name,
            "gstins": clean(row[offset + 2]),
            "officer_name": officer_name,
            "group_val": f"Group {group_raw}" if group_raw else None,
            "mode": clean(row[offset + 7]),
            "latest_status": clean(row[offset + 8]),
            "sio_email": clean(row[offset + 9]) if has_email else None,
        })

    # Group by FY, sort by date within each FY, assign sequential record_ids
    by_fy = defaultdict(list)
    for rec in raw_records:
        fy = fy_from_date(rec["date"])
        by_fy[fy].append(rec)

    assigned = []
    for fy in sorted(by_fy.keys()):
        recs = by_fy[fy]
        recs.sort(key=lambda r: (r["date"] is None, r["date"]))
        for seq, rec in enumerate(recs, 1):
            rec["record_id"] = f"NIR-{seq:03d}-{fy}"
            assigned.append(rec)

    # Batch-resolve sio emails → user IDs
    all_emails = [r.get("sio_email") for r in assigned if r.get("sio_email")]
    email_to_user_id = resolve_sio_emails(sb, workspace_id, all_emails) if not dry_run else {}
    if dry_run:
        print(f"  (dry run: skipping email→user_id resolution; {len(all_emails)} emails to resolve)\n")

    inserted = updated = skipped_count = 0

    for rec in assigned:
        sr_no = rec["sr_no"]
        sio_email = rec.get("sio_email")
        sio_user_id = email_to_user_id.get(sio_email) if sio_email else None

        payload = {
            "record_id": rec["record_id"],
            "legacy_non_ir_no": rec["legacy_non_ir_no"],
            "taxpayer_name": rec["taxpayer_name"],
            "gstins": rec["gstins"],
            "date_of_non_ir": rec["date"],
            "group": rec["group_val"],
            "sio_name": rec["officer_name"],
            "handling_io_sio": sio_user_id,
            "mode_of_initiation": rec["mode"],
            "latest_status": rec["latest_status"],
            "is_ir": False,
            "date_of_ir": None,
            "date_of_initiation": None,
        }
        payload = {k: v for k, v in payload.items() if v is not None or k in ("date_of_ir", "date_of_initiation", "record_id")}

        if dry_run:
            print(
                f"  [#{sr_no:03d}] taxpayer={rec['taxpayer_name']!r} | date={rec['date']}"
                f" | group={rec['group_val']!r} | sio={rec['officer_name']!r}"
                f" | sio_email={sio_email!r}"
                f" → {rec['record_id']}  legacy_non_ir_no={rec['legacy_non_ir_no']!r}"
            )

        result = upsert_non_ir_record(
            sb, workspace_id, sr_no, payload, skipped, log, dry_run
        )
        if result == "inserted":
            inserted += 1
        elif result == "updated":
            updated += 1
        else:
            skipped_count += 1

    print(
        f"\n[NON-IR]  {'Would insert' if dry_run else 'Inserted'}: {inserted}"
        f" | {'Would update' if dry_run else 'Updated'}: {updated}"
        f" | Skipped: {skipped_count}"
    )


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    args = sys.argv[1:]
    dry_run = "--dry-run" in args
    positional = [a for a in args if not a.startswith("--")]
    excel_path = os.path.abspath(positional[0] if positional else DEFAULT_EXCEL_PATH)

    if not os.path.exists(excel_path):
        raise SystemExit(f"Excel file not found: {excel_path}")

    if dry_run:
        print("*** DRY RUN — no changes will be written to the database ***\n")

    print(f"Loading: {excel_path}")
    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)

    sheet_name = next((s for s in SHEET_NAMES if s in wb.sheetnames), None)
    if sheet_name is None:
        raise SystemExit(f"No recognised sheet found. Available: {wb.sheetnames}")
    print(f"Sheet: {sheet_name!r}\n")

    sb = create_client(SUPABASE_URL, SERVICE_ROLE_KEY)

    res = (
        sb.table("votum_users")
        .select("workspace_id")
        .eq("email", WORKSPACE_OWNER_EMAIL)
        .limit(1)
        .execute()
    )
    if not res.data:
        raise SystemExit(f"No user found for {WORKSPACE_OWNER_EMAIL}")
    workspace_id = res.data[0]["workspace_id"]
    print(f"Workspace: {workspace_id}\n")

    skipped = []
    log = []
    process_sheet(wb[sheet_name], sb, workspace_id, skipped, log, dry_run)

    if log:
        with open(LOG_JSON, "w") as f:
            json.dump(log, f, indent=2)
        inserts = len([e for e in log if e["action"] == "insert"])
        updates = len([e for e in log if e["action"] == "update"])
        print(f"\nLog written to {LOG_JSON}  ({inserts} inserts, {updates} updates)")

    if skipped:
        if not dry_run:
            with open(SKIPPED_CSV, "w", newline="") as f:
                writer = csv.DictWriter(f, fieldnames=list(skipped[0].keys()))
                writer.writeheader()
                writer.writerows(skipped)
            print(f"\nSkipped {len(skipped)} rows — see {SKIPPED_CSV}")
        else:
            print(f"\nWould skip {len(skipped)} rows (errors during DB lookup).")
    else:
        print("\nNo rows skipped.")

    print("\nDry run complete — no data written." if dry_run else "\nDone.")


if __name__ == "__main__":
    main()
