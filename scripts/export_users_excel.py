"""
Export all seeded DGGI users (SIOs + DDs) with credentials to an Excel file.
Output: scripts/dggi_users.xlsx
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUTPUT = "scripts/dggi_users.xlsx"
DEFAULT_PASSWORD = "Dggi@1234"

GROUPS = ["Group A", "Group B", "Group C", "Group D", "Group E"]

def group_slug(g):
    return g.lower().replace(" ", "-")

# Build user list
users = []

# 25 SIOs
for group in GROUPS:
    for i in range(1, 6):
        slug = group_slug(group)
        users.append({
            "name": f"{group} SIO {i}",
            "email": f"{slug}-sio-{i}@dggi.gov.in",
            "role": "SIO",
            "group": group,
            "password": DEFAULT_PASSWORD,
        })

# 6 DDs
dd_entries = [
    ("Group A DD", "group-a-dd@dggi.gov.in", "Group A"),
    ("Group B DD", "group-b-dd@dggi.gov.in", "Group B"),
    ("Group C DD", "group-c-dd@dggi.gov.in", "Group C"),
    ("Group D DD", "group-d-dd@dggi.gov.in", "Group D"),
    ("Group E DD", "group-e-dd@dggi.gov.in", "Group E"),
    ("DD Unassigned", "dd-unassigned@dggi.gov.in", "—"),
]
for name, email, group in dd_entries:
    users.append({"name": name, "email": email, "role": "DD", "group": group, "password": DEFAULT_PASSWORD})

# ── Excel styling helpers ──────────────────────────────────────────────────
HEADER_FILL  = PatternFill("solid", fgColor="1F3864")  # dark navy
SIO_FILL     = PatternFill("solid", fgColor="DCE6F1")  # light blue
DD_FILL      = PatternFill("solid", fgColor="E2EFDA")  # light green
ALT_SIO_FILL = PatternFill("solid", fgColor="EEF4FB")
ALT_DD_FILL  = PatternFill("solid", fgColor="F0F7EC")
THIN = Side(style="thin", color="CCCCCC")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "DGGI Users"

# Title row
ws.merge_cells("A1:E1")
title_cell = ws["A1"]
title_cell.value = "DGGI Enforcement Portal — User Credentials"
title_cell.font = Font(name="Calibri", bold=True, size=14, color="FFFFFF")
title_cell.fill = PatternFill("solid", fgColor="1F3864")
title_cell.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[1].height = 28

# Header row
headers = ["#", "Name", "Email", "Role", "Group", "Password"]
ws.append([""] + headers)  # row 2, col A blank then headers in B–G
# Actually use columns A–F directly
ws.row_dimensions[2].height = 20
for col, header in enumerate(headers, start=1):
    cell = ws.cell(row=2, column=col, value=header)
    cell.font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="2E75B6")
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = BORDER

# Re-do title spanning correct columns
ws.merge_cells("A1:F1")

# Data rows
for i, u in enumerate(users, start=1):
    row = i + 2
    is_dd = u["role"] == "DD"
    base_fill = DD_FILL if is_dd else SIO_FILL
    alt_fill  = ALT_DD_FILL if is_dd else ALT_SIO_FILL
    fill = base_fill if i % 2 == 1 else alt_fill

    values = [i, u["name"], u["email"], u["role"], u["group"], u["password"]]
    for col, val in enumerate(values, start=1):
        cell = ws.cell(row=row, column=col, value=val)
        cell.fill = fill
        cell.border = BORDER
        cell.font = Font(name="Calibri", size=10)
        cell.alignment = Alignment(vertical="center")
        if col == 1:
            cell.alignment = Alignment(horizontal="center", vertical="center")
        if col == 4:  # Role badge center
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.font = Font(name="Calibri", size=10, bold=True,
                             color="155724" if is_dd else "0C3060")

# Column widths
col_widths = [4, 22, 34, 8, 12, 14]
for col, width in enumerate(col_widths, start=1):
    ws.column_dimensions[get_column_letter(col)].width = width

# Freeze header rows
ws.freeze_panes = "A3"

wb.save(OUTPUT)
print(f"Saved → {OUTPUT}  ({len(users)} users)")
